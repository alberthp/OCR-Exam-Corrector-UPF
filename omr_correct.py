#!/usr/bin/env python3
"""
OMR Exam Corrector - UPF Multi-Answer Test Auto-Correction
==========================================================

Processes scanned UPF exam PDFs and generates:
- Excel with student data, answers, and grades
- Annotated PDF for visual review

Author: Albert Hernansanz (with Claude)
Usage:
    python omr_correct.py exams.pdf students.csv answers.csv [options]
"""

import cv2
import numpy as np
import pandas as pd
import sys
import os
import time
import argparse
import traceback
from pdf2image import convert_from_path
from scipy.ndimage import uniform_filter1d
from scipy.signal import find_peaks
from PIL import Image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# =============================================
# CONFIGURATION CONSTANTS
# =============================================
DPI = 300
FILL_THRESHOLD_ANS = 0.10  # Threshold for answer detection
FILL_THRESHOLD_DIGIT = 0.35  # Minimum peak fill to accept a digit as marked
                              # Below this, the column is considered BLANK (no mark)
OPTION_LABELS = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']

# Reference bubble positions (calibrated at REFERENCE_DPI)
# All coordinates and pixel sizes are SCALED based on actual DPI before use.
REFERENCE_DPI = 300

REF_DNI_X      = [277, 327, 376, 427, 477, 527, 576, 626]
REF_CENTRE_X   = [827, 877, 927]
REF_ASSIGN_X   = [1027, 1077, 1127, 1177, 1226]
REF_PARCIAL_X  = [1325, 1375]
REF_PERMUT_X   = [1525]
REF_GRUP_X     = [1674, 1726]
REF_ID_X       = [1925, 1975, 2024, 2075, 2125, 2174, 2226, 2275]
REF_ANS_X      = [
    [275, 325, 375, 425, 475],
    [725, 775, 825, 875, 925],
    [1175, 1225, 1275, 1325, 1375],
    [1625, 1675, 1725, 1775, 1825],
    [2075, 2125, 2175, 2225, 2275],
]
# Bubble A x-position for each of the 5 answer columns (Q1-20, Q21-40, ..., Q81-100)
ANS_COL_A_X = [c[0] for c in REF_ANS_X]
# Spacing between option bubbles A->B->C... (calibrated at REFERENCE_DPI)
ANS_BUBBLE_SPACING = 50

# Bubble size half-dimensions (at REFERENCE_DPI)
BUBBLE_HALF_WIDTH = 12
BUBBLE_HALF_HEIGHT = 10

# Reference for x_offset cross-validation (at REFERENCE_DPI): on a calibration
# scan (page 2 corrected), the DNI box's left edge sits at x=320 when
# x_offset=77, so baseline DNI box x = 320 - 77 = 243.
REF_DNI_BOX_X = 243

# =============================================
# CORE IMAGE PROCESSING
# =============================================

def detect_id_boxes(img_bgr):
    """Detect the 7 labeled rectangular boxes in the ID header strip of the form.

    The physical form has seven red-bordered boxes across the top of the page,
    each containing a column of numbered bubbles: DNI (8 digits), CENTRE (3),
    ASSIGNATURA (5), PARCIAL (2), PERMUT (1), GRUP (2), and IDENTIFIER (8).
    These boxes are visually distinct from the answer area because they span
    roughly the top 50% of the page and are drawn with a thick red border.

    Returns a dict mapping field_name -> {'x', 'y', 'w', 'h'} for each
    detected box.  Used both as a QUALITY CHECK (all 7 found with consistent
    geometry -> perspective correction succeeded) and to position the orange
    outlines + value pills in the annotated PDF.
    """
    h, w = img_bgr.shape[:2]
    rf = img_bgr[:,:,2].astype(float)
    gf = img_bgr[:,:,1].astype(float)
    bf = img_bgr[:,:,0].astype(float)
    
    form_mask = np.uint8(
        (rf - np.maximum(gf, bf) > 3) & 
        (img_bgr.mean(axis=2) > 80) & 
        (img_bgr.mean(axis=2) < 240) & 
        (rf > 100)
    ) * 255
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))
    form_clean = cv2.morphologyEx(form_mask, cv2.MORPH_CLOSE, kernel, iterations=2)
    
    contours, _ = cv2.findContours(form_clean, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    
    candidates = []
    for c in contours:
        x, y, cw, ch = cv2.boundingRect(c)
        area = cv2.contourArea(c)
        bbox_area = cw * ch
        if bbox_area == 0: continue
        fr = area / bbox_area
        
        # Fill ratio (was 0.85, then 0.25) has kept getting lowered because it
        # measures how much of the box's printed content (border + bubble
        # outlines) the red-pixel mask picked up -- and that varies a lot
        # scan-to-scan (compression, focus, lighting) independent of whether
        # the box is genuinely there. Real data confirmed this: on a batch
        # where every page had exactly 7 candidate contours in this size
        # class (i.e. the geometry filter alone was already 100% precise,
        # zero spurious extras to reject), ASSIGNATURA still measured as low
        # as 0.232-0.234 on two pages -- just under the old 0.25 cutoff --
        # silently dropping it from validation and the annotated PDF outline.
        # CENTRE was measured as low as 0.301 in the same batch, one bad scan
        # from the same failure. 0.10 stays comfortably below every genuine
        # box observed while the size/position filters do the real work.
        if (550 < ch < 700 and 100 < cw < 480
            and y < h * 0.5 and fr > 0.10):
            candidates.append({'x': x, 'y': y, 'w': cw, 'h': ch})
    
    # Remove inner duplicates (smaller boxes inside larger ones)
    candidates.sort(key=lambda b: -b['w'] * b['h'])
    filtered = []
    for cand in candidates:
        is_inner = any(
            cand['x'] >= sel['x'] - 10 and 
            cand['x'] + cand['w'] <= sel['x'] + sel['w'] + 10 and
            cand['y'] >= sel['y'] - 10 and
            cand['y'] + cand['h'] <= sel['y'] + sel['h'] + 10
            for sel in filtered
        )
        if not is_inner:
            filtered.append(cand)
    
    # Identify each box by NEAREST expected reference x-position rather than
    # a naive left-to-right zip. A sequential zip silently shifts every field
    # after a gap into the wrong name when one box isn't detected (e.g. if
    # ASSIGNATURA's box is missing, a left-to-right zip would mislabel the
    # real PARCIAL box as ASSIGNATURA, the real PERMUT box as PARCIAL, etc.
    # -- which is exactly the bug that made a student's PARCIAL value show
    # up under the PERMUT box in the annotated PDF). Greedy nearest-match
    # instead leaves a field unassigned when its box is truly missing,
    # without disturbing the labels of the boxes that WERE found.
    field_ref_x = {
        'DNI': REF_DNI_X[0],
        'CENTRE': REF_CENTRE_X[0],
        'ASSIGNATURA': REF_ASSIGN_X[0],
        'PARCIAL': REF_PARCIAL_X[0],
        'PERMUT': REF_PERMUT_X[0],
        'GRUP': REF_GRUP_X[0],
        'IDENTIFIER': REF_ID_X[0],
    }
    # Max acceptable distance: smaller than the tightest gap between two
    # adjacent fields (PERMUT->GRUP, ~150px at REFERENCE_DPI), so a genuinely
    # missing box's name can never "steal" its neighbor's box.
    MAX_MATCH_DIST = 130

    pairs = []
    for name, ref_x in field_ref_x.items():
        for box in filtered:
            pairs.append((abs(box['x'] - ref_x), name, box))
    pairs.sort(key=lambda p: p[0])

    result = {}
    used_names = set()
    used_box_ids = set()
    for dist, name, box in pairs:
        if dist > MAX_MATCH_DIST:
            break
        box_id = id(box)
        if name in used_names or box_id in used_box_ids:
            continue
        result[name] = box
        used_names.add(name)
        used_box_ids.add(box_id)

    return result


def validate_perspective_correction(boxes):
    """Use the detected ID boxes to validate that perspective correction worked.

    quality_score is a rough 0-1 confidence figure: it starts at n_detected/7
    (fraction of expected boxes found) and is multiplied by 0.8 for each
    additional geometric inconsistency (e.g. width mismatch between DNI and
    IDENTIFIER, excessive vertical scatter).  A score of 1.0 means all 7 boxes
    were found with no geometric anomalies; below ~0.7 the page should be
    flagged for manual review.

    Returns dict with:
        - 'is_valid': bool - True if all 7 boxes detected with consistent geometry
        - 'n_detected': int - Number of boxes found
        - 'quality_score': float 0-1 - Overall quality of detection
        - 'issues': list of str - Description of any problems
    """
    issues = []
    expected_fields = ['DNI', 'CENTRE', 'ASSIGNATURA', 'PARCIAL', 'PERMUT', 'GRUP', 'IDENTIFIER']
    
    n_detected = sum(1 for f in expected_fields if f in boxes)
    
    if n_detected < 7:
        missing = [f for f in expected_fields if f not in boxes]
        issues.append(f"Missing boxes: {missing}")
    
    # Check that DNI and IDENTIFIER have similar dimensions (both 8 cols)
    if 'DNI' in boxes and 'IDENTIFIER' in boxes:
        dni_w = boxes['DNI']['w']
        id_w = boxes['IDENTIFIER']['w']
        if abs(dni_w - id_w) > 30:
            issues.append(f"DNI ({dni_w}px) and IDENTIFIER ({id_w}px) widths differ by {abs(dni_w-id_w)}px")
    
    # Check vertical alignment: all boxes should be in similar y-range
    if n_detected >= 2:
        ys = [boxes[f]['y'] for f in expected_fields if f in boxes]
        y_spread = max(ys) - min(ys)
        if y_spread > 50:
            issues.append(f"Boxes vertically misaligned (y spread = {y_spread}px)")
    
    quality_score = n_detected / 7.0
    if issues:
        quality_score *= 0.8  # Penalize issues
    
    return {
        'is_valid': n_detected == 7 and len(issues) == 0,
        'n_detected': n_detected,
        'quality_score': round(quality_score, 2),
        'issues': issues,
    }


def detect_form_rectangle(img_bgr):
    """Find the 4 corners of the answer area's red border rectangle.

    Isolates the red ink of the printed form border (high R, low G/B, mid
    brightness to exclude both white paper and black markers), morphologically
    closes small gaps, then picks the largest contour that spans at least 70%
    of page width and 40% of page height -- the outer border of the entire
    answer area, not one of the individual field boxes.

    Corner assignment uses the classic sum/difference trick on contour points:
      top-left     = smallest (x+y)    top-right    = largest (x-y)
      bottom-right = largest  (x+y)    bottom-left  = smallest (x-y)
    This is rotation-invariant for moderate skew angles (< ~20°).

    Returns dict with keys 'tl', 'tr', 'br', 'bl' (each a (x,y) ndarray),
    or None if no suitable contour is found.
    """
    h, w = img_bgr.shape[:2]
    rf = img_bgr[:,:,2].astype(float)
    gf = img_bgr[:,:,1].astype(float)
    bf = img_bgr[:,:,0].astype(float)

    form_mask = np.uint8(
        (rf - np.maximum(gf, bf) > 3) &
        (img_bgr.mean(axis=2) > 80) &
        (img_bgr.mean(axis=2) < 240) &
        (rf > 100)
    ) * 255

    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))
    form_clean = cv2.morphologyEx(form_mask, cv2.MORPH_CLOSE, kernel, iterations=2)

    contours, _ = cv2.findContours(form_clean, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    target = max(
        (c for c in contours
         if cv2.boundingRect(c)[2] > w*0.7 and cv2.boundingRect(c)[3] > h*0.4),
        key=cv2.contourArea, default=None
    )

    if target is None:
        return None

    pts = target.reshape(-1, 2)
    sums = pts[:, 0] + pts[:, 1]
    diffs = pts[:, 0] - pts[:, 1]
    tl = pts[np.argmin(sums)]
    br = pts[np.argmax(sums)]
    tr = pts[np.argmax(diffs)]
    bl = pts[np.argmin(diffs)]

    return {'tl': tl, 'tr': tr, 'br': br, 'bl': bl}


def correct_perspective(img_bgr, corners):
    """Apply a 4-point perspective warp to de-skew the scanned form.

    Computes a homography M that maps the detected (possibly skewed/rotated)
    form rectangle onto an axis-aligned rectangle of the same average width and
    height, anchored at the original top-left corner.  Anchoring at 'tl' keeps
    the corrected form in the same part of the canvas as the original, so the
    hardcoded pixel coordinates (REF_DNI_X, ANS_COL_A_X, etc.) remain valid
    without any additional translation.

    BORDER_REPLICATE fills the thin slivers outside the warped region with
    edge pixels, which is neutral (neither dark nor white) and avoids the
    edge-bright artifacts that BORDER_CONSTANT=white would cause near the
    left margin where markers are detected.

    Returns (corrected_bgr, M) where M is the 3x3 homography (kept for
    potential coordinate back-projection).
    """
    h, w = img_bgr.shape[:2]
    tl, tr, br, bl = corners['tl'], corners['tr'], corners['br'], corners['bl']

    # Average opposite-side lengths to handle mild trapezoidal distortion
    w_avg = int((np.linalg.norm(tr - tl) + np.linalg.norm(br - bl)) / 2)
    h_avg = int((np.linalg.norm(bl - tl) + np.linalg.norm(br - tr)) / 2)

    src = np.array([tl, tr, br, bl], dtype=np.float32)
    dst = np.array([
        tl,
        tl + [w_avg, 0],
        tl + [w_avg, h_avg],
        tl + [0, h_avg]
    ], dtype=np.float32)

    M = cv2.getPerspectiveTransform(src, dst)
    corrected = cv2.warpPerspective(
        img_bgr, M, (w, h),
        flags=cv2.INTER_LINEAR,
        borderMode=cv2.BORDER_REPLICATE
    )
    return corrected, M


def detect_markers(gray):
    """Detect the black alignment markers printed on the left margin of the form.

    The form has a column of solid black rectangles running down the left edge.
    Their y-positions encode the row grid: the first 10 markers align with the
    10 ID digit rows; then there is a visible gap; then 40 markers align with
    the 40 answer rows (20 questions x 2 rows each: answer + cancel).

    Detection strategy:
      1. Restrict to the leftmost 15% of the image (marker column only).
      2. Threshold with Otsu to handle scan-to-scan brightness variation.
      3. Project dark pixels onto the y-axis to get a row-intensity profile.
      4. Cluster bright peaks into individual marker centre rows.
      5. Find the largest spacing gap to split the run into id_rows / ans_rows.

    Returns (id_rows, ans_rows, median_spacing) as numpy arrays of y-pixel
    positions, or (None, None, None) if fewer than 20 markers are found.
    """
    h, w = gray.shape
    left = gray[:, :int(w*0.15)]
    # Adaptive (Otsu) threshold instead of a fixed <30 cutoff: scan-to-scan
    # brightness noise can push a marker's darkest pixel from ~28 to ~32,
    # which silently zeroes out detection under a hardcoded threshold even
    # though the marker is clearly distinguishable from the white background.
    _, dark_mask = cv2.threshold(left, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    dc = (dark_mask > 0).sum(axis=0)
    ds = uniform_filter1d(dc.astype(float), size=5)
    
    if ds.max() == 0:
        return None, None, None
    
    mc = np.where(ds > ds.max() * 0.2)[0]
    if len(mc) == 0:
        return None, None, None
    
    clusters = []
    s = mc[0]; p = mc[0]
    for c in mc[1:]:
        if c - p > 10:
            clusters.append((s, p)); s = c
        p = c
    clusters.append((s, p))
    best = max(clusters, key=lambda c: ds[c[0]:c[1]+1].sum())
    
    strip = gray[:, best[0]:best[1]+1]
    # Use MEAN intensity rather than MIN - markers are dark BLOCKS, so the mean
    # of pixels in a row crossing a marker is low. This is more robust to
    # internal noise/compression artifacts within markers (which can give the
    # interior a slightly lighter shade than the borders).
    rm = strip.mean(axis=1)
    dr = np.where(rm < 150)[0]
    
    if len(dr) == 0:
        return None, None, None
    
    # Marker minimum height: at least 5 px tall (filters out single-pixel noise
    # while accepting both thick markers from domestic scanners and the typical
    # 20-25 px markers from professional scanners like Ricoh)
    MIN_MARKER_HEIGHT = 5
    
    markers = []
    s = dr[0]; p = dr[0]
    for r in dr[1:]:
        if r - p > 3:
            if p - s >= MIN_MARKER_HEIGHT:
                markers.append((s + p) // 2)
            s = r
        p = r
    if p - s >= MIN_MARKER_HEIGHT:
        markers.append((s + p) // 2)
    
    if len(markers) < 20:
        return None, None, None
    
    sps = [markers[i+1] - markers[i] for i in range(len(markers)-1)]
    median_sp = sorted(sps)[len(sps)//2]
    gap_candidates = [(i, sps[i]) for i in range(len(sps)) if sps[i] > median_sp*2]
    
    if not gap_candidates:
        return None, None, None
    
    best_gap = min(gap_candidates, 
                   key=lambda x: abs(x[0]+1-10) + abs(len(markers)-x[0]-1-40))[0]
    
    id_rows = np.array(markers[:best_gap+1])[-10:]
    ans_rows = np.array(markers[best_gap+1:])[:40]
    
    if len(ans_rows) < 40 and len(ans_rows) >= 30:
        sp = np.median(np.diff(ans_rows))
        while len(ans_rows) < 40:
            ans_rows = np.append(ans_rows, ans_rows[-1] + sp)
    
    return id_rows, ans_rows, median_sp


def find_x_offset(img_bgr, id_rows):
    """Find the horizontal shift of the form relative to the hardcoded reference coordinates.

    The reference bubble x-positions (REF_DNI_X, REF_CENTRE_X, etc.) were
    calibrated on a specific scan where the form's left edge happened to land
    at a particular x.  Every real scan may shift the form slightly left or
    right (due to paper feed variation or the scanner's paper guide).
    x_offset is added to every REF_*_X value before reading a bubble, so a
    positive offset means the form printed further to the right than the
    reference scan.

    Strategy: look for a group of 6-10 small red blobs (the ID bubble column
    outlines) in the horizontal band spanned by id_rows, find the one group
    whose x-positions cluster with the expected DNI column spacing, and derive
    the shift from the leftmost blob vs. the calibrated DNI first-column x.
    Falls back to 77 (the empirically observed median offset) on failure.
    """
    h, w = img_bgr.shape[:2]
    rf = img_bgr[:,:,2].astype(float)
    gf = img_bgr[:,:,1].astype(float)
    bf = img_bgr[:,:,0].astype(float)
    
    fm = np.uint8(
        (rf - np.maximum(gf, bf) > 3) & 
        (img_bgr.mean(axis=2) > 80) & 
        (img_bgr.mean(axis=2) < 240) & 
        (rf > 100)
    ) * 255
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
    fc = cv2.morphologyEx(fm, cv2.MORPH_CLOSE, kernel, iterations=1)
    
    y1 = int(id_rows[0]) - 15
    y2 = int(id_rows[-1]) + 15
    cnt, _ = cv2.findContours(fc[y1:y2, :], cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    all_cx = sorted([
        x + cv2.boundingRect(c)[2]//2 
        for c in cnt 
        for x in [cv2.boundingRect(c)[0]] 
        if 12 < cv2.boundingRect(c)[2] < 45 
        and 10 < cv2.boundingRect(c)[3] < 35
    ])
    
    if not all_cx:
        return 77
    
    sections = [[all_cx[0]]]
    for x in all_cx[1:]:
        if x - sections[-1][-1] > 60:
            sections.append([x])
        else:
            sections[-1].append(x)
    
    for sec in sections:
        if 6 <= len(sec) <= 10:
            offset = min(sec) - 277
            if -50 < offset < 300:
                return offset
    
    return 77


def calibrate_color_thresholds(img_bgr, id_y_range=None):
    """Calibrate per-page brightness threshold by sampling anchor regions.

    Different scanner models and paper stocks produce different brightness
    levels for the same ink.  A fixed threshold (e.g. <160 = dark) would
    either miss light pencil marks on high-contrast scanners or include form
    elements on low-contrast scanners.  Instead, this function samples two
    known-reference regions -- the left-margin marker strip (definitely dark)
    and the top-right blank corner (definitely paper white) -- and places the
    threshold 55% of the way from dark to paper.  The 55% bias toward the dark
    side ensures lightly-filled pencil bubbles (closer to the paper end of the
    scale) are still caught.

    Returns a dict with:
        'brightness_threshold': int   -- pixels below this are treated as marks
        'dark_reference': float       -- mean brightness of sampled dark pixels
        'paper_reference': float      -- mean brightness of sampled paper pixels
    """
    h, w = img_bgr.shape[:2]
    
    # Sample known dark pixels: left margin where markers are
    left_strip = img_bgr[:, :int(w*0.08), :]
    gray_left = cv2.cvtColor(left_strip, cv2.COLOR_BGR2GRAY)
    very_dark_mask = gray_left < 50
    
    if very_dark_mask.sum() > 100:
        dark_pixels = left_strip[very_dark_mask]
        dark_brightness = dark_pixels.mean(axis=1).mean()
    else:
        dark_brightness = 30  # fallback
    
    # Sample paper background: pixels far from any feature
    # Use the top-right corner area which is usually clean
    bg_strip = img_bgr[int(h*0.02):int(h*0.05), int(w*0.5):int(w*0.95), :]
    gray_bg = cv2.cvtColor(bg_strip, cv2.COLOR_BGR2GRAY)
    paper_mask = gray_bg > 200
    if paper_mask.sum() > 100:
        paper_brightness = gray_bg[paper_mask].mean()
    else:
        paper_brightness = 240  # fallback
    
    # Adaptive brightness threshold: midpoint between dark and paper, biased toward dark
    # This accommodates light pencil marks while excluding light form pink
    brightness_threshold = int(dark_brightness + (paper_brightness - dark_brightness) * 0.55)
    brightness_threshold = max(120, min(200, brightness_threshold))
    
    return {
        'brightness_threshold': brightness_threshold,
        'dark_reference': dark_brightness,
        'paper_reference': paper_brightness,
    }


def make_student_mask(img_bgr, color_cal=None):
    """Build a binary mask that isolates student marks and excludes the printed form.

    The mask is 255 where a pixel is likely a student mark (pencil or blue/dark
    ink) and 0 everywhere else (paper white, red form lines, red bubble outlines).
    A pixel is accepted if it is either:
      - Generally dark (brightness < threshold) AND not strongly red-dominant
        (this catches pencil and dark-ink fills)
      - More blue than red by at least 10 counts AND not too bright
        (this catches students who use blue ballpoint or felt-tip pens)

    The resulting mask is the sole input to read_bubble_fill() -- no pixel-level
    color information flows past this point, only fill ratios.
    """
    bf = img_bgr[:,:,0].astype(float)
    gf = img_bgr[:,:,1].astype(float)
    rf = img_bgr[:,:,2].astype(float)
    brightness = img_bgr.mean(axis=2)
    red_dom = rf - (gf + bf) / 2
    
    # Use adaptive threshold if provided, else use safe default
    bt = color_cal['brightness_threshold'] if color_cal else 170
    
    # red_dom<40 (not 20): dark ink can pick up a slight red tint where it
    # overlaps the printed pink bubble outline (verified on a real mark that
    # was 89% excluded at <20 despite being unambiguously dark by brightness).
    # Printed form elements are already excluded by the brightness check, so
    # this stays safe against false positives on unmarked bubbles.
    mask = np.uint8(
        ((brightness < bt) & (red_dom < 40)) |  # Dark marks (pencil/dark ink)
        ((bf > rf + 10) & (brightness < bt + 20))    # Blue ink (more permissive on brightness)
    ) * 255
    return mask


def read_bubble_fill(mask, cx, cy):
    """Return the fraction of mask pixels that are set (255) in a bubble's ROI.

    The ROI is a rectangle of size (2*BUBBLE_HALF_WIDTH) x (2*BUBBLE_HALF_HEIGHT)
    centred on (cx, cy).  The fill ratio is the count of set pixels divided by
    total pixels; an empty/unmarked bubble returns ~0.01-0.04 (noise from form
    lines), while a clearly filled bubble returns 0.15-0.80 depending on how
    heavily the student marked it.  FILL_THRESHOLD_ANS / FILL_THRESHOLD_DIGIT
    are set relative to this scale.
    """
    y1 = max(0, cy - BUBBLE_HALF_HEIGHT)
    y2 = min(mask.shape[0], cy + BUBBLE_HALF_HEIGHT)
    x1 = max(0, cx - BUBBLE_HALF_WIDTH)
    x2 = min(mask.shape[1], cx + BUBBLE_HALF_WIDTH)
    roi = mask[y1:y2, x1:x2]
    return roi.sum() / (roi.size * 255.0) if roi.size > 0 else 0.0


def compute_adaptive_digit_threshold(peak_fills, default=0.35,
                                       min_thresh=0.20, max_thresh=0.50):
    """Find the natural gap separating noise from real marks in this page's fill distribution.

    The fill values across all ID columns on a page are assumed to be bimodal:
    a low cluster near 0 (empty bubbles -- noise from printed outlines) and a
    higher cluster (filled bubbles -- student marks).  The threshold is placed
    at the midpoint of the largest gap found between consecutive sorted fills,
    provided the gap is at least 0.10 wide.

    On very lightly-marked pages (faint pencil) the gap may be narrow, causing
    the function to fall back to `default`.  The clamp [min_thresh, max_thresh]
    prevents degenerate thresholds from slipping into either cluster.

    Args:
        peak_fills: list of per-column peak fill ratios sampled from the page
        default: fallback threshold if no clear gap is found (width < 0.10)
        min_thresh: lower bound on returned threshold (noise floor guard)
        max_thresh: upper bound on returned threshold (mark acceptance guard)

    Returns:
        float: threshold for this page; a fill above this is accepted as a mark
    """
    if not peak_fills:
        return default
    
    sorted_fills = sorted(peak_fills)
    
    # Case 1: All values low -> empty page (no marks)
    if max(sorted_fills) < min_thresh:
        return 0.99  # Reject everything
    
    # Case 2: All values high -> well-marked page
    if min(sorted_fills) > max_thresh:
        return min_thresh  # Accept everything
    
    # Case 3: Find the largest gap in the noise/signal boundary range
    best_gap = 0
    best_threshold = default
    
    for i in range(len(sorted_fills) - 1):
        a = sorted_fills[i]
        b = sorted_fills[i + 1]
        if a < max_thresh and b > min_thresh:
            gap = b - a
            if gap > best_gap:
                best_gap = gap
                best_threshold = (a + b) / 2
    
    # If no significant gap, use default
    if best_gap < 0.10:
        return default
    
    return max(min_thresh, min(max_thresh, best_threshold))


def read_digit_marker_anchored(mask, cx, id_rows, median_sp, w,
                                  fill_threshold=None):
    """Detect which digit (0-9) is filled in a single ID bubble column.

    "Marker-anchored" means the result is only accepted if the detected ink
    peak aligns with a known marker row (within 70% of the median spacing).
    This prevents spurious reads from form lines, shadow artifacts, or ink
    bleed in the gap between the ID and answer sections.

    Strategy:
      1. Extract a vertical strip of the student mask covering the full ID row range.
      2. Project horizontally (mean across the strip's width) to get a 1-D fill profile.
      3. Find the tallest peak in that profile.
      4. Accept only if: peak fill >= fill_threshold AND the peak y aligns with
         one of id_rows (distance < 0.7 * median_sp).
      5. Return the 0-based index of the matching id_row as the digit value.

    Args:
        mask: binary student mask (output of make_student_mask)
        cx: x centre of the bubble column (already offset-corrected)
        id_rows: array of y-pixel positions for digit rows 0-9
        median_sp: median row-to-row spacing in pixels (from detect_markers)
        w: image width (used for boundary clamping)
        fill_threshold: minimum peak fill to accept as a mark;
                        defaults to FILL_THRESHOLD_DIGIT if None

    Returns:
        (digit, fill): digit is the row index (0-9) or None; fill is the peak
        fill ratio (useful for adaptive threshold calibration even on None returns)
    """
    if fill_threshold is None:
        fill_threshold = FILL_THRESHOLD_DIGIT
    
    y_margin = int(median_sp * 0.8)
    y_top = int(id_rows[0]) - y_margin
    y_bot = int(id_rows[-1]) + y_margin
    x1 = max(0, cx - BUBBLE_HALF_WIDTH)
    x2 = min(w, cx + BUBBLE_HALF_WIDTH)
    col_strip = mask[y_top:y_bot, x1:x2]
    
    if col_strip.size == 0:
        return None, 0
    
    profile = uniform_filter1d(col_strip.mean(axis=1) / 255.0, size=5)
    peaks, props = find_peaks(
        profile, height=0.03, 
        distance=int(median_sp * 0.4),
        prominence=0.015
    )
    
    if len(peaks) == 0:
        return None, float(profile.max())
    
    best_peak = peaks[np.argmax(props['peak_heights'])]
    peak_y_abs = best_peak + y_top
    peak_fill = float(profile[best_peak])
    
    # Reject as noise if the strongest "mark" is too weak
    if peak_fill < fill_threshold:
        return None, peak_fill
    
    dists = np.abs(id_rows - peak_y_abs)
    if dists.min() < median_sp * 0.7:
        return int(np.argmin(dists)), peak_fill
    
    return None, peak_fill


def get_column_peak_fill(mask, cx, id_rows, median_sp, w):
    """Return the strongest fill peak in a column strip without any threshold or digit check.

    Used to build the list of raw fill values that compute_adaptive_digit_threshold()
    uses to locate the noise/signal gap.  Applying the threshold here would be circular
    (the threshold hasn't been computed yet).
    """
    y_margin = int(median_sp * 0.8)
    y_top = int(id_rows[0]) - y_margin
    y_bot = int(id_rows[-1]) + y_margin
    x1 = max(0, cx - BUBBLE_HALF_WIDTH)
    x2 = min(w, cx + BUBBLE_HALF_WIDTH)
    col_strip = mask[y_top:y_bot, x1:x2]
    if col_strip.size == 0:
        return 0.0
    profile = uniform_filter1d(col_strip.mean(axis=1) / 255.0, size=5)
    return float(profile.max())


# =============================================
# DECODING LOGIC
# =============================================

def decode_identifier(digits):
    """Decode the student's 6-digit U-number from a raw 8-bubble read.

    The IDENTIFIER field has 8 bubble columns, but the U-number is only 6
    digits.  Students may bubble the 6 digits in any 6 consecutive positions,
    padding with leading or trailing zeros for the remaining 2.  This function
    handles the ambiguity:

      * 8 non-null digits, first two zeros  -> strip leading '00', take digits 2-7
      * 8 non-null digits, last two zeros   -> take digits 0-5
      * 8 non-null digits, both ambiguous   -> return 'AMBIGUOUS' with both candidates
      * exactly 6 non-null digits           -> return as-is ('OK')
      * 7 non-null digits with a leading/trailing 0 -> strip the pad, return 'OK_PADDED'
      * < 6 non-null digits                 -> return 'INCOMPLETE'

    Returns (u_number_str, status_str).  Status is used downstream to colour-code
    the IDENTIFIER cell in the Excel output and the annotated PDF header.
    """
    non_null = [(i, d) for i, d in enumerate(digits) if d is not None]
    
    if len(non_null) == 0:
        return None, 'MISSING'
    
    if len(non_null) == 8:
        all_str = ''.join(str(d) for d in digits)
        if digits[0] == 0 and digits[1] == 0:
            return all_str[2:], 'OK_PADDED'
        if digits[6] == 0 and digits[7] == 0:
            return all_str[:6], 'OK_PADDED'
        return all_str[:6] + '|' + all_str[2:], 'AMBIGUOUS'
    
    if len(non_null) == 6:
        return ''.join(str(d) for _, d in non_null), 'OK'
    
    if len(non_null) == 7:
        # Which of the 8 bubble positions is the one that didn't register.
        # This -- not merely "the boundary digit happens to be 0" -- is
        # what actually tells leading/trailing-pad apart from a missed
        # REAL digit that happens to sit next to a genuine pad zero. The
        # previous check (`non_null[-1][1] == 0`) could fire even when the
        # missing bubble was one of the 6 real digits and a genuine pad
        # zero elsewhere just happened to be last in the list, silently
        # returning a wrong-but-plausible 6-digit ID tagged OK_PADDED
        # (confirmed: digits=[2,3,4,None,6,7,0,0] -> old code returned
        # "234670", dropping the missing digit instead of the pad zero).
        missing_idx = next(i for i, d in enumerate(digits) if d is None)
        s = ''.join(str(d) for _, d in non_null)
        if missing_idx == 0 and digits[1] == 0:
            return s[1:], 'OK_PADDED'
        if missing_idx == 7 and digits[6] == 0:
            return s[:6], 'OK_PADDED'
        # The missing bubble is one of the 6 real digits (or the pad
        # position that DID register isn't actually 0) -- the dropped
        # digit can't be recovered, so surface the raw 7-digit read
        # instead of confidently guessing a specific wrong 6-digit ID.
        return s, 'WARNING'
    
    if len(non_null) < 6:
        return ''.join(str(d) for _, d in non_null), 'INCOMPLETE'
    
    return ''.join(str(d) for _, d in non_null), 'WARNING'


def decode_grup(digits):
    """Decode GRUP field (1 or 2 digits, flexible position)."""
    non_null = [d for d in digits if d is not None]
    if len(non_null) == 0:
        return None, 'BLANK'
    return ''.join(str(d) for d in non_null), 'OK'


# =============================================
# ANSWER READING + SCORING
# =============================================

def read_answers(mask, x_offset, ans_rows, num_questions, num_options=5,
                 fill_threshold=None):
    """Read all student answers, applying the form's two-row cancel mechanism.

    Each question occupies TWO consecutive marker rows in ans_rows:
      - Row 2k   (ans_row):  the student marks their chosen option(s) here
      - Row 2k+1 (can_row):  the student can cancel a mark by filling the same
                              option column in this row

    Cancellation logic:
      - net_marked = bubbles filled in ans_row BUT NOT in can_row
        (a cancelled bubble is excluded from the answer set)
      - If only cancel-row bubbles are present with no ans-row marks -> ONLY_CANCEL
        (form filled out in wrong order; treated as no answer)
      - If ans-row marks are fully negated by cancel-row marks -> FULLY_CANCELLED

    Status values:
      'OK'              -- normal (may be blank or marked)
      'ONLY_CANCEL'     -- cancel row has marks but answer row does not
      'FULLY_CANCELLED' -- all answer-row marks were cancelled

    Args:
        fill_threshold: minimum fill ratio to count as marked; uses
                        FILL_THRESHOLD_ANS if None
    """
    if fill_threshold is None:
        fill_threshold = FILL_THRESHOLD_ANS
    
    OPTS = OPTION_LABELS[:num_options]
    answers = {}
    
    for q in range(1, num_questions + 1):
        col_idx = (q - 1) // 20
        q_in_col = (q - 1) % 20
        ans_row = q_in_col * 2
        can_row = q_in_col * 2 + 1
        
        if can_row >= len(ans_rows):
            break
        
        # Generate bubble x-positions dynamically based on num_options
        # Each column has bubbles at: A_x, A_x + spacing, A_x + 2*spacing, ...
        a_x = ANS_COL_A_X[col_idx]
        col_x = [a_x + i * ANS_BUBBLE_SPACING + x_offset for i in range(num_options)]
        ay = int(ans_rows[ans_row])
        cy = int(ans_rows[can_row])
        
        ans_fills = np.array([read_bubble_fill(mask, x, ay) for x in col_x])
        can_fills = np.array([read_bubble_fill(mask, x, cy) for x in col_x])
        
        ans_marked = ans_fills > fill_threshold
        can_marked = can_fills > fill_threshold
        net_marked = ans_marked & ~can_marked
        
        marked_options = [OPTS[i] for i in range(num_options) if net_marked[i]]
        
        if not any(ans_marked) and not any(can_marked):
            status = 'OK'; marks_set = set()
        elif marked_options:
            status = 'OK'; marks_set = set(marked_options)
        elif any(can_marked) and not any(ans_marked):
            status = 'ONLY_CANCEL'; marks_set = set()
        else:
            status = 'FULLY_CANCELLED'; marks_set = set()
        
        answers[q] = {
            'marks': marks_set,
            'status': status,
            'ans_fills': ans_fills.tolist(),
            'can_fills': can_fills.tolist(),
            'ans_y': ay, 'can_y': cy, 'col_x': col_x,
        }
    
    return answers


def score_question(student_marks, correct_answers, num_options):
    """Compute the partial-credit score for one question.

    UPF multi-answer partial-credit formula:
      - Marking a correct option:   +1 / |CORRECT|
      - Marking an incorrect option: -1 / (num_options - |CORRECT|)
      - Not marking an option: 0
      - Total is clamped to [0, 1] (no negative score per question)

    Rationale: a student who marks all options scores 0 (positives and
    negatives cancel exactly), which prevents gambling on unknown questions.
    A student who marks exactly the correct set scores 1.0.

    Args:
        student_marks: set of option labels the student marked (e.g. {'A', 'C'})
        correct_answers: set of correct option labels (e.g. {'A', 'B'})
        num_options: total number of options available for this question

    Returns:
        float in [0, 1]: the student's score for this question
    """
    good = len(correct_answers)
    if good == 0 or good == num_options:
        return 0.0  # Invalid question definition
    
    pos_weight = 1.0 / good
    neg_weight = 1.0 / (num_options - good)
    
    score = 0.0
    for opt in student_marks:
        if opt in correct_answers:
            score += pos_weight
        else:
            score -= neg_weight
    
    return max(0.0, score)


# =============================================
# MAIN PIPELINE
# =============================================

def detect_page_orientation(img_bgr):
    """Detect which edge of the page contains the marker column.

    The black alignment markers form a column at one edge of the form.
    In portrait orientation (correct), markers are on the LEFT edge.
    In other orientations, they're on TOP/RIGHT/BOTTOM.

    Returns: 'left' (correct), 'top', 'right', or 'bottom'
    """
    # Rotations chosen to match auto_rotate_to_portrait's mapping: the edge
    # named here is brought onto the LEFT margin by the given rotation.
    rotations = {
        'left': img_bgr,
        'top': cv2.rotate(img_bgr, cv2.ROTATE_90_COUNTERCLOCKWISE),
        'right': cv2.rotate(img_bgr, cv2.ROTATE_180),
        'bottom': cv2.rotate(img_bgr, cv2.ROTATE_90_CLOCKWISE),
    }

    # For each candidate edge, check whether rotating it to the left margin
    # yields a genuine periodic marker pattern (not just raw dark-pixel count,
    # which is easily skewed by heavily-inked answer bubbles landing in the
    # sampled strip). detect_markers() validates spacing regularity, so reuse
    # it here instead of a separate brightness heuristic.
    best_edge = None
    best_n_markers = 0
    for edge, rotated in rotations.items():
        gray = cv2.cvtColor(rotated, cv2.COLOR_BGR2GRAY)
        id_rows, ans_rows, _ = detect_markers(gray)
        if id_rows is None:
            continue
        n_markers = len(id_rows) + len(ans_rows)
        if n_markers > best_n_markers:
            best_n_markers = n_markers
            best_edge = edge

    if best_edge is not None:
        return best_edge

    # Fallback: no rotation produced a valid marker pattern (e.g. severely
    # degraded scan). Use the old raw dark-pixel-count heuristic as a guess.
    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    h, w = gray.shape
    strip_size = 0.15
    strips = {
        'left':   gray[:, :int(w * strip_size)],
        'right':  gray[:, int(w * (1 - strip_size)):],
        'top':    gray[:int(h * strip_size), :],
        'bottom': gray[int(h * (1 - strip_size)):, :],
    }
    scores = {name: int((strip < 30).sum()) for name, strip in strips.items()}
    return max(scores, key=scores.get)


def auto_rotate_to_portrait(img_bgr):
    """Automatically rotate the image so the marker column is on the LEFT edge.
    
    Returns (rotated_image, rotation_applied_in_degrees)
    """
    edge = detect_page_orientation(img_bgr)
    
    # Map edge to rotation needed to bring markers to LEFT
    # If markers are at:
    #   left   -> already correct (0°)
    #   top    -> rotate 90° counter-clockwise (or 270° clockwise)
    #   right  -> rotate 180°
    #   bottom -> rotate 90° clockwise
    if edge == 'left':
        return img_bgr, 0
    elif edge == 'top':
        return cv2.rotate(img_bgr, cv2.ROTATE_90_COUNTERCLOCKWISE), -90
    elif edge == 'right':
        return cv2.rotate(img_bgr, cv2.ROTATE_180), 180
    elif edge == 'bottom':
        return cv2.rotate(img_bgr, cv2.ROTATE_90_CLOCKWISE), 90
    return img_bgr, 0


def process_page(img_pil, page_num, num_questions, num_options=5, source_dpi=300):
    """Run the full OMR pipeline for one scanned exam page.

    Pipeline stages (each stage feeds into the next):
      0a. Auto-rotate   -- bring the left-margin marker column to the left edge
      0b. Normalize DPI -- rescale to REFERENCE_DPI so all constants apply
      1.  Corner detect -- find the 4 corners of the red form border
      2.  Perspective   -- warp the image to remove skew/rotation
      3.  Marker detect -- locate the 10 ID rows and 40 answer rows by y-position
      4.  X-offset      -- measure how far the form shifted horizontally
      5a. Color calibrate-- determine per-page brightness threshold
      5b. Student mask  -- isolate student ink from the red form
      5c-d. Adaptive threshold -- find the noise/signal gap in fill values
      5e. Box detection -- validate perspective via ID field box geometry
      5f. Offset XVal   -- cross-validate x_offset using detected DNI box
      6.  ID fields     -- read DNI, CENTRE, ASSIGNATURA, PARCIAL, PERMUT, GRUP, IDENTIFIER
      7.  Answer threshold -- compute adaptive threshold for answer bubbles
      8.  Answers       -- read all question answers with cancel-row logic

    Args:
        img_pil: PIL image of the scanned page
        page_num: page number (1-indexed, for error reporting)
        num_questions: number of questions to grade
        num_options: number of options per question (default 5)
        source_dpi: DPI the PDF was rendered at; image is rescaled to
                    REFERENCE_DPI internally so all hardcoded coordinates stay valid

    Returns:
        (result_dict, corrected_bgr): result carries all detected data plus
        private '_corrected', '_mask', '_id_rows', '_ans_rows', '_median_sp'
        arrays for downstream annotation/caching.  Returns (result, None) on
        early failure (CORNER_ERROR or MARKER_ERROR).
    """
    img = np.array(img_pil)
    img_bgr = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)
    
    result = {
        'page': page_num,
        'status': 'OK',
        'errors': [],
        'source_dpi': source_dpi,
    }
    
    # Step 0a: Auto-rotate to portrait orientation (markers on left edge)
    img_bgr, rotation_applied = auto_rotate_to_portrait(img_bgr)
    result['rotation_applied'] = rotation_applied
    
    # Step 0b: Normalize image to REFERENCE_DPI so all hardcoded coordinates work
    # This is much cleaner than scaling every constant.
    if source_dpi != REFERENCE_DPI:
        scale = REFERENCE_DPI / source_dpi
        h_new = int(img_bgr.shape[0] * scale)
        w_new = int(img_bgr.shape[1] * scale)
        img_bgr = cv2.resize(img_bgr, (w_new, h_new), interpolation=cv2.INTER_AREA)
        result['resized_to_ref_dpi'] = True
    
    # Step 1: Find form rectangle corners
    corners = detect_form_rectangle(img_bgr)
    if corners is None:
        result['status'] = 'CORNER_ERROR'
        result['errors'].append('Could not detect form rectangle')
        return result, None
    
    # Step 2: Apply perspective correction
    corrected, M = correct_perspective(img_bgr, corners)
    
    # Step 3: Detect markers on corrected image
    gray = cv2.cvtColor(corrected, cv2.COLOR_BGR2GRAY)
    id_rows, ans_rows, median_sp = detect_markers(gray)
    if id_rows is None:
        result['status'] = 'MARKER_ERROR'
        result['errors'].append('Could not detect markers after correction')
        return result, None
    
    # Step 4: Find x-offset
    x_offset = find_x_offset(corrected, id_rows)
    result['x_offset'] = x_offset

    # Step 5a: Calibrate color thresholds for this page (adaptive RGB)
    color_cal = calibrate_color_thresholds(corrected)
    result['color_calibration'] = color_cal
    
    # Step 5b: Make student mask with adaptive thresholds
    mask = make_student_mask(corrected, color_cal=color_cal)
    h, w = corrected.shape[:2]
    
    # Step 5c: Sample peak fills from all ID columns to learn the noise/signal boundary
    sample_cols = (REF_DNI_X + REF_CENTRE_X + REF_ASSIGN_X + 
                   REF_PARCIAL_X + REF_PERMUT_X + REF_GRUP_X + REF_ID_X)
    sample_fills = []
    for ref in sample_cols:
        cx = ref + x_offset
        fill = get_column_peak_fill(mask, cx, id_rows, median_sp, w)
        sample_fills.append(fill)
    
    # Step 5d: Compute adaptive digit threshold for this page
    adaptive_threshold = compute_adaptive_digit_threshold(sample_fills)
    result['digit_threshold'] = round(adaptive_threshold, 3)
    
    # Step 5e: QUALITY VALIDATION via ID box detection
    detected_boxes = detect_id_boxes(corrected)
    validation = validate_perspective_correction(detected_boxes)
    result['quality_score'] = validation['quality_score']
    result['n_boxes_detected'] = validation['n_detected']
    result['validation_issues'] = validation['issues']
    if not validation['is_valid']:
        result['errors'].extend(validation['issues'])
    result['_detected_boxes'] = detected_boxes
    
    # Step 5f: Cross-validate x_offset using detected DNI box.
    # If actual DNI box position differs significantly, prefer the box-based offset.
    if 'DNI' in detected_boxes:
        dni_box_x = detected_boxes['DNI']['x']
        expected_box_x = REF_DNI_BOX_X + x_offset
        box_offset_drift = dni_box_x - expected_box_x
        result['box_offset_drift'] = box_offset_drift
        
        # Only correct if drift is SIGNIFICANT (>30px = more than one bubble width)
        # Small drifts are within scan noise tolerance
        if abs(box_offset_drift) > 30:
            corrected_x_offset = dni_box_x - REF_DNI_BOX_X
            result['x_offset_original'] = x_offset
            result['x_offset_corrected_from_box'] = True
            x_offset = corrected_x_offset
            result['x_offset'] = x_offset
    
    # Step 6: Read ID fields with adaptive threshold
    fields = {
        'DNI': REF_DNI_X,
        'CENTRE': REF_CENTRE_X,
        'ASSIGNATURA': REF_ASSIGN_X,
        'PARCIAL': REF_PARCIAL_X,
        'PERMUT': REF_PERMUT_X,
        'GRUP': REF_GRUP_X,
        'IDENTIFIER': REF_ID_X,
    }
    
    raw_fields = {}
    for fname, cols in fields.items():
        digits = []
        for c in cols:
            d, _ = read_digit_marker_anchored(mask, c + x_offset, id_rows, median_sp, w,
                                                fill_threshold=adaptive_threshold)
            digits.append(d)
        raw_fields[fname] = digits
    
    # Decode
    result['dni'] = ''.join(str(d) if d is not None else '_' for d in raw_fields['DNI'])
    result['centre'] = ''.join(str(d) if d is not None else '_' for d in raw_fields['CENTRE'])
    result['assignatura'] = ''.join(str(d) if d is not None else '_' for d in raw_fields['ASSIGNATURA'])
    
    parcial_nn = [d for d in raw_fields['PARCIAL'] if d is not None]
    result['parcial'] = ''.join(str(d) for d in parcial_nn) if parcial_nn else None
    
    result['permut'] = raw_fields['PERMUT'][0]
    
    grup_val, _ = decode_grup(raw_fields['GRUP'])
    result['grup'] = grup_val
    
    u_val, u_status = decode_identifier(raw_fields['IDENTIFIER'])
    result['u_number'] = u_val
    result['u_status'] = u_status
    result['identifier_raw'] = ''.join(str(d) if d is not None else '_' for d in raw_fields['IDENTIFIER'])
    
    # Step 7: Compute adaptive answer threshold by sampling all answer bubbles
    answer_sample_fills = []
    for q in range(1, num_questions + 1):
        col_idx = (q - 1) // 20
        q_in_col = (q - 1) % 20
        ans_row = q_in_col * 2
        if ans_row >= len(ans_rows):
            break
        a_x = ANS_COL_A_X[col_idx]
        col_x = [a_x + i * ANS_BUBBLE_SPACING + x_offset for i in range(num_options)]
        ay = int(ans_rows[ans_row])
        for x in col_x:
            answer_sample_fills.append(read_bubble_fill(mask, x, ay))
    
    answer_threshold = compute_adaptive_digit_threshold(
        answer_sample_fills, default=FILL_THRESHOLD_ANS, 
        min_thresh=0.08, max_thresh=0.30
    )
    result['answer_threshold'] = round(answer_threshold, 3)
    
    # Step 8: Read answers with adaptive threshold
    answers = read_answers(mask, x_offset, ans_rows, num_questions, num_options,
                            fill_threshold=answer_threshold)
    result['answers'] = answers

    result['_corrected'] = corrected
    result['_mask'] = mask
    result['_id_rows'] = id_rows
    result['_ans_rows'] = ans_rows
    result['_median_sp'] = median_sp
    
    return result, corrected


# =============================================
# OUTPUT GENERATION
# =============================================

def _write_results_sheet(ws, results_list, student_lookup, correct_answers,
                          num_questions, num_options=5):
    """Populate one worksheet with a results table for a single permutation.

    Column layout (left to right):
      Cols 1-18 (fixed, see `fixed_headers` below -- kept as the single
      source of truth so this comment can't drift out of sync with the code
      again): Page, Status, U_Number, U_Status, Name, Surname1, Surname2,
      Email, DNI, PARCIAL, PERMUT, GRUP, GRUP_Check, N_Answered, Grade,
      Grade_10, ID_Problem, Manual_Edit
      Then for each question Q (repeated num_questions times):
        (num_options) option columns  -- 1 if student marked that option, else 0
        1 Score column                -- partial-credit score for this question

    Row layout:
      Row 1: merged "Q{n}" header spanning each question's (num_options+1) columns
      Row 2: column sub-headers (A/B/C/D/E + Score, repeated per question)
      Row 3: CORRECT ANSWERS reference row (1/0 per option, score = 1.0 or blank)
      Row 4+: one data row per scanned page

    An empty correct_answers dict is passed for the No_Perm_Detected sheet, which
    causes all Score cells to remain blank rather than scoring against a wrong key.
    """
    hdr_font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    hdr_fill = PatternFill('solid', fgColor='2C3E50')
    ok_fill = PatternFill('solid', fgColor='D5F5E3')
    warn_fill = PatternFill('solid', fgColor='FCF3CF')
    err_fill = PatternFill('solid', fgColor='FADBD8')
    border = Border(*[Side(style='thin')]*4)
    
    fixed_headers = ['Page', 'Status', 'U_Number', 'U_Status',
                     'Name', 'Surname1', 'Surname2', 'Email',
                     'DNI', 'PARCIAL', 'PERMUT', 'GRUP', 'GRUP_Check',
                     'N_Answered', 'Grade', 'Grade_10', 'ID_Problem', 'Manual_Edit']
    n_fixed = len(fixed_headers)
    OPTS_USED = OPTION_LABELS[:num_options]
    
    # ===== Row 1: Question group headers (merged cells per question) =====
    # Fixed columns (1..n_fixed): empty in row 1, will hold the actual header in row 2
    for ci in range(1, n_fixed + 1):
        cell = ws.cell(row=1, column=ci, value='')
        cell.fill = hdr_fill
        cell.border = border
    
    # For each question: merge (num_options + 1) cells with "Qn" label
    for q in range(1, num_questions + 1):
        q_col_start = n_fixed + (q - 1) * (num_options + 1) + 1
        q_col_end = q_col_start + num_options  # includes Score column
        
        # Merge the cells across all options + score
        start_letter = openpyxl.utils.get_column_letter(q_col_start)
        end_letter = openpyxl.utils.get_column_letter(q_col_end)
        ws.merge_cells(f'{start_letter}1:{end_letter}1')
        
        cell = ws.cell(row=1, column=q_col_start, value=f'Q{q}')
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # ===== Row 2: Detailed column headers =====
    # Fixed columns
    for ci, h in enumerate(fixed_headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Option letters + Score per question
    for q in range(1, num_questions + 1):
        q_col_start = n_fixed + (q - 1) * (num_options + 1) + 1
        for oi, opt in enumerate(OPTS_USED):
            cell = ws.cell(row=2, column=q_col_start + oi, value=opt)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        # Score column header
        score_cell = ws.cell(row=2, column=q_col_start + num_options, value='Score')
        score_cell.font = hdr_font
        score_cell.fill = hdr_fill
        score_cell.alignment = Alignment(horizontal='center')
        score_cell.border = border
    
    # Row 3: CORRECT ANSWERS reference row
    correct_hdr_fill = PatternFill('solid', fgColor='3498DB')
    correct_hdr_font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    
    label_cell = ws.cell(row=3, column=1, value='CORRECT')
    label_cell.font = correct_hdr_font
    label_cell.fill = correct_hdr_fill
    label_cell.alignment = Alignment(horizontal='center')
    label_cell.border = border
    # Span the label across the rest of the fixed columns
    for ci in range(2, n_fixed + 1):
        c = ws.cell(row=3, column=ci, value='')
        c.fill = correct_hdr_fill
        c.border = border
    
    # Fill the correct answer pattern for each question
    for q in range(1, num_questions + 1):
        q_col_start = n_fixed + (q - 1) * (num_options + 1) + 1
        correct_set = correct_answers.get(q, set())
        for oi, opt in enumerate(OPTS_USED):
            cell = ws.cell(row=3, column=q_col_start + oi,
                          value=1 if opt in correct_set else 0)
            cell.fill = correct_hdr_fill
            cell.font = correct_hdr_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        # Max possible score for this question
        score_cell = ws.cell(row=3, column=q_col_start + num_options,
                            value=1.0 if correct_set else '')
        score_cell.fill = correct_hdr_fill
        score_cell.font = correct_hdr_font
        score_cell.alignment = Alignment(horizontal='center')
        score_cell.border = border
    
    for ri, r in enumerate(results_list, 4):
        page = r.get('page')
        u_num = r.get('u_number', '')
        u_clean = str(u_num).split('|')[0] if u_num else ''
        u_status = r.get('u_status', '')
        
        # Match student
        nom, c1, c2, email = '', '', '', ''
        id_problem = 'OK'
        matched_student = None
        if u_clean in student_lookup:
            matched_student = student_lookup[u_clean]
            nom = matched_student.get('Nom', '')
            c1 = matched_student.get('Cognom1', '')
            c2 = matched_student.get('Cognom2', '') or ''
            email = matched_student.get('Email', '') or ''
        elif u_clean:
            id_problem = 'UNUMBER_NO_MATCH'
        else:
            id_problem = u_status if u_status != 'OK' else 'UNUMBER_MISSING'

        ws.cell(row=ri, column=1, value=page)
        ws.cell(row=ri, column=2, value=r.get('status', ''))
        ws.cell(row=ri, column=3, value=u_clean)
        cell_us = ws.cell(row=ri, column=4, value=u_status)
        if u_status == 'OK' or u_status == 'OK_PADDED':
            cell_us.fill = ok_fill
        elif u_status in ['AMBIGUOUS', 'INCOMPLETE', 'WARNING']:
            cell_us.fill = warn_fill
        elif u_status == 'MISSING':
            cell_us.fill = err_fill

        ws.cell(row=ri, column=5, value=nom)
        ws.cell(row=ri, column=6, value=c1)
        ws.cell(row=ri, column=7, value=c2)
        ws.cell(row=ri, column=8, value=email)
        ws.cell(row=ri, column=9, value=r.get('dni', ''))
        ws.cell(row=ri, column=10, value=r.get('parcial', ''))
        permut_val = r.get('permut')
        ws.cell(row=ri, column=11, value='' if permut_val is None else str(permut_val))

        # GRUP: may have been backfilled/cross-checked against the student
        # roster's theory group by backfill_and_validate_groups() -- see
        # 'grup_check' ('OK' / 'FROM_ROSTER' / 'MISMATCH (...)' / '').
        grup_cell = ws.cell(row=ri, column=12, value=r.get('grup', ''))
        grup_check = r.get('grup_check', '')
        if grup_check == 'OK':
            grup_cell.fill = ok_fill
        elif grup_check == 'FROM_ROSTER':
            grup_cell.fill = warn_fill
        elif grup_check.startswith('MISMATCH'):
            grup_cell.fill = err_fill
        ws.cell(row=ri, column=13, value=grup_check)

        # Compute grades
        answers = r.get('answers', {})
        n_answered = sum(1 for a in answers.values() if a.get('marks'))
        ws.cell(row=ri, column=14, value=n_answered)
        
        total_score = 0.0
        max_score = 0.0
        for q in range(1, num_questions + 1):
            correct_set = correct_answers.get(q, set())
            student_marks = answers.get(q, {}).get('marks', set())
            
            if correct_set:
                q_score = score_question(student_marks, correct_set, num_options)
                total_score += q_score
                max_score += 1.0
                
                # Annotate answer status for the annotated PDF's green/yellow/red
                # bubble coloring. This is a separate, simplified 3-way category
                # (exact match / partial-no-wrong-marks / anything else) -- NOT
                # the same as q_score above, which uses score_question()'s
                # continuous partial-credit formula for the actual grade.
                if student_marks == correct_set:
                    answers.setdefault(q, {})['score_info'] = {'all_correct': True, 'partial': False}
                elif student_marks & correct_set and not (student_marks - correct_set):
                    answers.setdefault(q, {})['score_info'] = {'all_correct': False, 'partial': True}
                else:
                    answers.setdefault(q, {})['score_info'] = {'all_correct': False, 'partial': False}
            else:
                q_score = None
        
        ws.cell(row=ri, column=15, value=round(total_score, 3))
        grade_10 = round((total_score / max_score) * 10, 2) if max_score > 0 else 0
        ws.cell(row=ri, column=16, value=grade_10)

        cell_id = ws.cell(row=ri, column=17, value=id_problem)
        if id_problem == 'OK':
            cell_id.fill = ok_fill
        else:
            cell_id.fill = err_fill

        cell_manual = ws.cell(row=ri, column=18, value='Y' if r.get('_manual_edit') else 'N')
        if r.get('_manual_edit'):
            cell_manual.fill = warn_fill

        # Q columns: one column per option (Q1_A, Q1_B, ...) with 1/0
        for q in range(1, num_questions + 1):
            q_data = answers.get(q, {})
            marks = q_data.get('marks', set())
            
            # Column offset: after the n_fixed fixed columns, each question takes (num_options + 1) columns
            q_col_start = n_fixed + (q - 1) * (num_options + 1) + 1
            
            # Determine fill color for this question
            if q in correct_answers:
                correct_set = correct_answers[q]
                if marks == correct_set:
                    fill = ok_fill
                elif marks & correct_set and not (marks - correct_set):
                    fill = warn_fill
                elif marks:
                    fill = err_fill
                else:
                    fill = None
            else:
                fill = None
            
            # Write 1/0 for each option
            for oi, opt in enumerate(OPTS_USED):
                cell = ws.cell(row=ri, column=q_col_start + oi, 
                              value=1 if opt in marks else 0)
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
                if fill is not None and opt in marks:
                    cell.fill = fill
            
            # Q_Score column at the end of the question block
            score_col = q_col_start + num_options
            if q in correct_answers:
                q_score = score_question(marks, correct_answers[q], num_options)
                score_cell = ws.cell(row=ri, column=score_col, value=round(q_score, 3))
                score_cell.alignment = Alignment(horizontal='center')
                score_cell.border = border
            else:
                ws.cell(row=ri, column=score_col, value='').border = border
        
        for ci in range(1, n_fixed + 1):
            ws.cell(row=ri, column=ci).border = border
    
    # Column widths (indexed to fixed_headers' order, not hardcoded letters,
    # so inserting/removing a fixed column doesn't silently misalign widths)
    widths_by_header = {
        'Page': 6, 'Status': 10, 'U_Number': 10, 'U_Status': 12,
        'Name': 14, 'Surname1': 14, 'Surname2': 14, 'Email': 26, 'DNI': 12,
        'PARCIAL': 8, 'PERMUT': 8, 'GRUP': 8, 'GRUP_Check': 20,
        'N_Answered': 8, 'Grade': 8, 'Grade_10': 8, 'ID_Problem': 18,
        'Manual_Edit': 10,
    }
    for ci, h in enumerate(fixed_headers, 1):
        w = widths_by_header.get(h)
        if w:
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
    
    # Narrow columns for option markers (1/0) and slightly wider for scores
    for q in range(1, num_questions + 1):
        q_col_start = n_fixed + (q - 1) * (num_options + 1) + 1
        for oi in range(num_options):
            col_letter = openpyxl.utils.get_column_letter(q_col_start + oi)
            ws.column_dimensions[col_letter].width = 4
        score_col_letter = openpyxl.utils.get_column_letter(q_col_start + num_options)
        ws.column_dimensions[score_col_letter].width = 7
    
    ws.freeze_panes = f"{openpyxl.utils.get_column_letter(n_fixed + 1)}4"


def _normalize_group_value(value):
    """'01' and '1' both mean theory group 1. The GRUP bubble field has two
    digit columns (padded, e.g. "01"/"02"), while a roster-derived
    TheoryGroup is a bare single digit ("1"/"2") -- comparing them as raw
    strings made every genuinely-matching page register as a false
    MISMATCH. Strips the leading zero via int() when the value parses as
    one; falls back to the stripped string for anything else.
    """
    s = str(value).strip()
    try:
        return str(int(s))
    except (TypeError, ValueError):
        return s


def _normalize_perm_value(value):
    """Perm labels are small integers, but pandas silently upcasts the
    *entire* Perm column to float64 if even one cell elsewhere in it is
    blank/NaN -- str(0.0) is "0.0", not "0". The scanned PERMUT bubble
    always produces a clean Python int (str(0) == "0"), so a plain str()
    cast on a float-upcast answer-key column can never match it again: every
    page for that permutation silently routes to the No_Perm_Detected sheet
    and grades against an empty key, with no error anywhere. Routes through
    float() first (unlike _normalize_group_value) specifically to collapse
    that "0.0" case back to "0"; falls back to the stripped string for
    non-numeric Perm labels (e.g. "A"/"B").
    """
    s = str(value).strip()
    try:
        return str(int(float(s)))
    except (TypeError, ValueError):
        return s


def backfill_and_validate_groups(all_results, students_df):
    """Cross-check each page's scanned GRUP (theory group) against the
    student roster, and backfill it when the bubble couldn't be read.

    The exam bubble sheet has a single GRUP field; UPF's Moodle "participants"
    roster export additionally encodes a theory group + seminar subgroup
    (e.g. Grups="201-7"), from which load_students() derives a 'TheoryGroup'
    column (the leading digit -- there's no bubble field for the seminar
    subgroup, so that part is dropped). If students_df has no such column
    (older/simpler roster formats), this is a no-op.

    For each result matched to a roster student by U-number:
      - scanned GRUP blank, roster has a theory group -> backfill r['grup']
        from the roster; r['grup_source'] = 'roster', r['grup_check'] =
        'FROM_ROSTER' (both output paths use this to mark it as
        roster-derived rather than scanned, per the "indicate the origin"
        requirement).
      - scanned GRUP matches roster                    -> r['grup_check'] = 'OK'
      - scanned GRUP conflicts with roster              -> r['grup_check'] =
        'MISMATCH (scan=X, roster=Y)'. NOT auto-corrected: a misread bubble
        and a stale/wrong roster entry are both plausible, so this is left
        for a human to resolve rather than silently overwritten either way.
      - no roster theory group, or student unmatched     -> r['grup_check'] = ''

    Mutates each result dict in place. Returns None.
    """
    if students_df is None or 'TheoryGroup' not in students_df.columns:
        return

    roster = {}
    for _, row in students_df.iterrows():
        u = str(row.get('U_number', '')).strip().upper().replace('U', '')
        tg = row.get('TheoryGroup')
        if u and tg and str(tg) not in ('', 'nan', 'None'):
            roster[u] = str(tg).strip()

    for r in all_results:
        u_clean = str(r.get('u_number', '') or '').split('|')[0]
        roster_grup = roster.get(u_clean)
        r['grup_source'] = 'scan'
        if not roster_grup:
            r['grup_check'] = ''
            continue
        scanned_grup = r.get('grup')
        if scanned_grup in (None, ''):
            r['grup'] = roster_grup
            r['grup_source'] = 'roster'
            r['grup_check'] = 'FROM_ROSTER'
        elif _normalize_group_value(scanned_grup) == _normalize_group_value(roster_grup):
            r['grup_check'] = 'OK'
        else:
            r['grup_check'] = f'MISMATCH (scan={scanned_grup}, roster={roster_grup})'


def _write_group_sheet(ws, all_results, correct_answers_by_perm, group_value,
                        num_questions, num_options):
    """Populate a flat DNI/U_Number/Score roster for one theory group (T1/T2).

    Unlike the per-permutation sheets, this cuts across PERMUT: a theory
    group's students may have sat either exam permutation, so each row is
    graded against whichever permutation that specific page detected,
    using the same total_score/max_score/grade_10 formula as
    _write_results_sheet() (so a student's number here always matches their
    Grade_10 on their "Perm N" sheet).

    Matches on the (possibly roster-backfilled) GRUP field via
    _normalize_group_value(), so a scanned "01" and a roster "1" are treated
    as the same group -- see backfill_and_validate_groups().

    A page with no GRUP at all (unread and no roster to backfill from) is
    excluded from every group sheet rather than guessed into one.
    """
    hdr_font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    hdr_fill = PatternFill('solid', fgColor='2C3E50')
    border = Border(*[Side(style='thin')]*4)

    headers = ['DNI', 'U_Number', 'Score']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    ri = 2
    for r in all_results:
        grup_val = r.get('grup')
        if grup_val in (None, '') or _normalize_group_value(grup_val) != group_value:
            continue

        permut_val = r.get('permut')
        correct_by_q = (correct_answers_by_perm.get(str(permut_val))
                         if permut_val is not None else None)
        answers = r.get('answers', {})
        total_score = 0.0
        max_score = 0.0
        if correct_by_q:
            for q in range(1, num_questions + 1):
                correct_set = correct_by_q.get(q, set())
                if not correct_set:
                    continue
                student_marks = answers.get(q, {}).get('marks', set())
                total_score += score_question(student_marks, correct_set, num_options)
                max_score += 1.0
        grade_10 = round((total_score / max_score) * 10, 2) if max_score > 0 else ''

        ws.cell(row=ri, column=1, value=r.get('dni', ''))
        ws.cell(row=ri, column=2, value=str(r.get('u_number') or '').split('|')[0])
        ws.cell(row=ri, column=3, value=grade_10)
        for ci in range(1, 4):
            ws.cell(row=ri, column=ci).border = border
        ri += 1

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.freeze_panes = 'A2'


def write_excel(all_results, students_df, correct_answers_by_perm, output_path,
                 num_questions, num_options=5):
    """Generate an Excel workbook with one results sheet per exam permutation.

    Sheet structure:
      "Perm 0", "Perm 1", ... -- one sheet per known permutation; each page
                                  is graded against that permutation's answer key
      "No_Perm_Detected"      -- pages where the PERMUT bubble was unreadable;
                                  answer columns shown but Score left blank
      "T1", "T2"              -- flat DNI/U_Number/Score roster per theory
                                  group (see _write_group_sheet()), cutting
                                  across PERMUT since GRUP and PERMUT are
                                  independent fields
      "Summary"               -- totals: pages, U-matches, per-permutation counts

    Each scanned page is routed to the sheet matching its OWN detected PERMUT
    value and graded with that permutation's key.  The routing is a pure string
    match (str(permut_val) == perm_key), so permutation labels must be consistent
    between the answer file and what the bubble reader outputs.

    correct_answers_by_perm: dict returned by load_correct_answers(), mapping
    perm_key (str) -> {question_num: set_of_correct_option_letters}.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Build student lookup by U-number (shared across all sheets + summary)
    student_lookup = {}
    if students_df is not None:
        for _, row in students_df.iterrows():
            u = str(row.get('U_number', '')).strip().upper().replace('U', '')
            if u:
                student_lookup[u] = row

    # Sort permutation keys numerically where possible (e.g. '0','1','2'),
    # falling back to lexicographic order for non-numeric perm labels.
    def _perm_sort_key(p):
        try:
            return (0, int(p))
        except (ValueError, TypeError):
            return (1, str(p))
    known_perms = sorted(correct_answers_by_perm.keys(), key=_perm_sort_key)

    # Route each page to the sheet matching its own detected PERMUT
    groups = {p: [] for p in known_perms}
    no_perm_group = []
    for r in all_results:
        permut_val = r.get('permut')
        permut_str = _normalize_perm_value(permut_val) if permut_val is not None else None
        if permut_str is not None and permut_str in groups:
            groups[permut_str].append(r)
        else:
            no_perm_group.append(r)

    for p in known_perms:
        ws = wb.create_sheet(f"Perm {p}")
        _write_results_sheet(ws, groups[p], student_lookup, correct_answers_by_perm[p],
                              num_questions, num_options)

    ws_noperm = wb.create_sheet("No_Perm_Detected")
    _write_results_sheet(ws_noperm, no_perm_group, student_lookup, {},
                          num_questions, num_options)

    # ===== Theory-group roster sheets (T1/T2) =====
    # Flat DNI/U-Number/Score view per theory group, independent of exam
    # permutation -- GRUP (backfill_and_validate_groups() may have filled or
    # cross-checked it against the roster) is orthogonal to PERMUT, so a T1
    # student can appear on either the "Perm 1" or "Perm 2" sheet above.
    for group_value in ('1', '2'):
        ws_t = wb.create_sheet(f"T{group_value}")
        _write_group_sheet(ws_t, all_results, correct_answers_by_perm,
                            group_value, num_questions, num_options)

    # ===== Summary sheet =====
    ws2 = wb.create_sheet("Summary")
    ws2['A1'] = 'OMR Processing Summary'
    ws2['A1'].font = Font(bold=True, size=14)

    summary_data = [
        ('Total pages', len(all_results)),
        ('Pages with answers', sum(1 for r in all_results if r.get('answers'))),
        ('U-numbers matched', sum(1 for r in all_results
                                   if str(r.get('u_number','')).split('|')[0] in student_lookup)),
        ('U-numbers unmatched', sum(1 for r in all_results
                                     if str(r.get('u_number','')).split('|')[0]
                                     and str(r.get('u_number','')).split('|')[0] not in student_lookup)),
        ('Num questions', num_questions),
        ('Num options', num_options),
    ]
    for p in known_perms:
        summary_data.append((f'Pages with Perm {p}', len(groups[p])))
    summary_data.append(('Pages with no Perm detected', len(no_perm_group)))

    for ri, (k, v) in enumerate(summary_data, 3):
        ws2.cell(row=ri, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=ri, column=2, value=v)

    wb.save(output_path)
    return output_path


def _build_student_name_lookup(students_df):
    """U-number -> full display name, used for the annotated PDF header."""
    student_lookup = {}
    if students_df is not None:
        for _, row in students_df.iterrows():
            u = str(row.get('U_number', '')).strip().upper().replace('U', '')
            if u:
                nom = str(row.get('Nom', '')).strip()
                c1 = str(row.get('Cognom1', '')).strip()
                c2 = str(row.get('Cognom2', '') or '').strip()
                if c2.lower() == 'nan':
                    c2 = ''
                full_name = f"{nom} {c1} {c2}".strip()
                student_lookup[u] = full_name
    return student_lookup


def _draw_plain_scan_page(c, r, page_w_pt, page_h_pt):
    """Draw just the perspective-corrected scan, centered on the page, with
    no header/footer/overlays -- the "as scanned" copy of a page for a
    formal grade-review request, before any grading annotation is added.

    Uses the same image-fit math as _draw_annotated_page() (margin, aspect
    fit) so the scan lines up identically across both pages of the export.
    """
    import io
    from reportlab.lib.utils import ImageReader

    corr = r.get('_corrected')
    if corr is None:
        return

    img_h, img_w = corr.shape[:2]
    target_h = 1700
    if img_h > target_h:
        scale = target_h / img_h
        new_w = int(img_w * scale)
        img_for_pdf = cv2.resize(corr, (new_w, target_h), interpolation=cv2.INTER_AREA)
    else:
        img_for_pdf = corr

    rgb = cv2.cvtColor(img_for_pdf, cv2.COLOR_BGR2RGB)
    pil_img = Image.fromarray(rgb)

    margin = 8
    avail_w = page_w_pt - 2 * margin
    avail_h = page_h_pt - 2 * margin
    img_aspect = img_for_pdf.shape[1] / img_for_pdf.shape[0]
    page_aspect = avail_w / avail_h
    if img_aspect > page_aspect:
        draw_w = avail_w
        draw_h = avail_w / img_aspect
    else:
        draw_h = avail_h
        draw_w = avail_h * img_aspect
    img_x_pt = (page_w_pt - draw_w) / 2
    img_y_pt = (page_h_pt - draw_h) / 2

    img_buf = io.BytesIO()
    pil_img.save(img_buf, format='JPEG', quality=90)
    img_buf.seek(0)
    c.drawImage(ImageReader(img_buf), img_x_pt, img_y_pt, width=draw_w, height=draw_h)
    c.showPage()


def export_student_review_pdf(result, output_path, students_df=None,
                                correct_answers_by_perm=None):
    """Write a 2-page PDF for a single student's grade-review request.

    Page 1 is the raw scanned page with no annotation layer (what the
    student actually filled in); page 2 is the same page with the full
    grading annotation layer (as shown in annotated_review.pdf), plus --
    when correct_answers_by_perm is given -- the expected-answer overlay
    and colour legend, so a reviewer handling the request has everything
    needed without opening the full batch PDF or the review GUI.
    """
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4

    PAGE_W_PT, PAGE_H_PT = A4
    student_lookup = _build_student_name_lookup(students_df)

    c = canvas.Canvas(output_path, pagesize=A4)
    _draw_plain_scan_page(c, result, PAGE_W_PT, PAGE_H_PT)
    _draw_annotated_page(c, result, student_lookup, PAGE_W_PT, PAGE_H_PT,
                          correct_answers_by_perm=correct_answers_by_perm)
    c.save()
    return output_path


def _draw_annotated_page(c, r, student_lookup, page_w_pt, page_h_pt,
                          correct_answers_by_perm=None):
    """Draw one exam page with vector annotation overlays onto a reportlab canvas.

    Layout (bottom-to-top in PDF coordinate space):
      [footer 36pt]  -- detected answer summary + any validation warnings
      [image area]   -- rasterized corrected scan, aspect-ratio-fitted to A4
      [header 60pt]  -- status badge, U-number, name, quality metrics

    Coordinate system: PDF origin is BOTTOM-LEFT; image pixels are TOP-LEFT.
    The helper px_to_pt() converts image (x, y) -> PDF (x_pt, y_pt), flipping
    the y-axis.  All vector overlays (bubble rectangles, box outlines, text
    pills) are drawn in PDF point space on top of the raster background.

    Color conventions for answer bubble overlays:
      GREEN  -- student marks exactly matching the correct set
      YELLOW -- partial match (some correct marks, no wrong marks)
      RED    -- wrong or mixed marks
      BLUE   -- cancel-row marks (student crossed out a bubble)
      PURPLE -- marks added or removed by a human reviewer after OMR

    Shared by write_annotated_pdf (full batch) and patch_annotated_pdf_page
    (single-page splice after a manual correction) so a page can be redrawn
    without regenerating the entire PDF.

    correct_answers_by_perm: optional (as returned by load_correct_answers());
    when given, additionally draws a blue diagonal-slash "expected answer"
    overlay over every bubble the key marks correct for this page's
    permutation, and a colour legend in the footer -- used by
    export_student_review_pdf() for a single-student review copy. Regular
    batch/patch callers omit it, so annotated_review.pdf is unaffected.
    """
    from reportlab.lib.colors import Color
    import io

    PAGE_W_PT, PAGE_H_PT = page_w_pt, page_h_pt

    # Colors (RGB 0-1)
    GREEN = Color(0, 0.74, 0)
    YELLOW = Color(1, 0.82, 0)
    RED = Color(0.86, 0, 0)
    BLUE = Color(0, 0.47, 0.86)
    ORANGE = Color(1, 0.39, 0)
    PURPLE = Color(0.58, 0.0, 0.83)  # marks/fields a reviewer added or changed by hand
    TEAL = Color(0, 0.55, 0.55)  # GRUP value backfilled from the roster, not scanned
    GRAY_DARK = Color(0.15, 0.15, 0.15)
    GRAY = Color(0.4, 0.4, 0.4)
    GRAY_LIGHT = Color(0.94, 0.94, 0.94)
    BG_GREEN = Color(0.78, 0.92, 0.78)
    BG_YELLOW = Color(1.0, 0.95, 0.7)
    BG_RED = Color(1, 0.78, 0.78)
    PILL_BG = Color(1, 1, 0.78)
    WHITE = Color(1, 1, 1)

    corr = r.get('_corrected')
    if corr is None:
        return

    img_h, img_w = corr.shape[:2]

    # The header/footer bars below are opaque rectangles painted OVER the
    # full-bleed image, not layout boxes the image is fit inside -- so
    # anything that grows the footer (the legend strip) must also shrink the
    # image's available height by the same amount, or it silently starts
    # covering real bubble rows instead of the sheet's blank bottom margin.
    show_legend = bool(correct_answers_by_perm)
    legend_extra_h = 14 if show_legend else 0

    # ========================================================
    # 1. Background image: the corrected page (raster)
    # ========================================================
    # Convert image to JPEG bytes (compressed; we don't need 300 DPI for screen)
    # Resize to ~150 DPI to keep file size manageable
    target_h = 1700
    if img_h > target_h:
        scale = target_h / img_h
        new_w = int(img_w * scale)
        img_for_pdf = cv2.resize(corr, (new_w, target_h), interpolation=cv2.INTER_AREA)
    else:
        img_for_pdf = corr

    # Convert BGR to RGB for PIL
    rgb = cv2.cvtColor(img_for_pdf, cv2.COLOR_BGR2RGB)
    pil_img = Image.fromarray(rgb)

    # Compute the image's position on the PDF page (centered, full A4)
    # We work in PDF coordinates (points, origin bottom-left)
    # Page area available: A4 minus margins. Sized exactly as the regular
    # (no-legend) pipeline -- legend_extra_h is applied purely as a vertical
    # shift below, not by shrinking the fit, so this size/aspect math is
    # identical either way.
    margin = 8  # pts
    avail_w = PAGE_W_PT - 2 * margin
    avail_h = PAGE_H_PT - 2 * margin

    # Fit image within A4 maintaining aspect ratio
    img_aspect = img_for_pdf.shape[1] / img_for_pdf.shape[0]
    page_aspect = avail_w / avail_h
    if img_aspect > page_aspect:
        draw_w = avail_w
        draw_h = avail_w / img_aspect
    else:
        draw_h = avail_h
        draw_w = avail_h * img_aspect

    # Position image centered, then shifted up by legend_extra_h so the
    # enlarged footer hides exactly the same footer-height's worth of image
    # as the regular 36pt footer does -- not 14pt more. The image simply
    # extends a little further up behind the opaque header at the top,
    # which already covers far more than that either way.
    img_x_pt = (PAGE_W_PT - draw_w) / 2
    img_y_pt = (PAGE_H_PT - draw_h) / 2 + legend_extra_h

    # Save image to in-memory buffer and draw on PDF
    img_buf = io.BytesIO()
    pil_img.save(img_buf, format='JPEG', quality=80)
    img_buf.seek(0)
    from reportlab.lib.utils import ImageReader
    c.drawImage(ImageReader(img_buf), img_x_pt, img_y_pt,
                width=draw_w, height=draw_h)
    
    # ========================================================
    # 2. Coordinate transformer: image pixels -> PDF points
    # ========================================================
    # IMPORTANT: PDF origin is BOTTOM-LEFT, image origin is TOP-LEFT
    scale_x = draw_w / img_w  # image pixels -> PDF points
    scale_y = draw_h / img_h
    
    def px_to_pt(x, y):
        """Convert (x, y) in image-pixel space to PDF point coordinates."""
        return (img_x_pt + x * scale_x,
                img_y_pt + draw_h - y * scale_y)
    
    def px_w(w):  # width in image pixels -> width in points
        return w * scale_x
    
    def px_h(h):
        return h * scale_y
    
    # ========================================================
    # 3. Annotations - all as VECTOR graphics
    # ========================================================
    
    # --- HEADER (top of page, drawn as filled rectangle + text) ---
    page_num = r.get('page', '?')
    u = r.get('u_number', '?')
    u_status = r.get('u_status', '?')
    dni = r.get('dni', '?')
    quality = r.get('quality_score', 0)
    n_boxes = r.get('n_boxes_detected', 0)
    n_answered = sum(1 for a in r.get('answers', {}).values() if a.get('marks'))
    digit_thr = r.get('digit_threshold', '?')
    ans_thr = r.get('answer_threshold', '?')
    
    u_clean = str(u).split('|')[0] if u else ''
    expected_name = student_lookup.get(u_clean, '') if u_clean else ''
    
    # Determine header status
    if quality >= 1.0 and n_answered > 0 and u and u_status in ('OK', 'OK_PADDED'):
        header_color = BG_GREEN
        status_text = "OK"
        status_text_color = GREEN
    elif quality >= 0.7 and (n_answered > 0 or u):
        header_color = BG_YELLOW
        status_text = "REVIEW"
        status_text_color = ORANGE
    else:
        header_color = BG_RED
        status_text = "MANUAL CHECK"
        status_text_color = RED
    
    # Draw header bar (top of page, above image). Taller on manually-edited
    # pages to leave room for the extra legend line below the badge.
    hdr_h_pt = 70 if r.get('_manual_edit') else 60
    hdr_y_pt = PAGE_H_PT - hdr_h_pt
    c.setFillColor(header_color)
    c.rect(0, hdr_y_pt, PAGE_W_PT, hdr_h_pt, fill=1, stroke=0)
    c.setStrokeColor(GRAY)
    c.setLineWidth(0.5)
    c.line(0, hdr_y_pt, PAGE_W_PT, hdr_y_pt)

    # Line 1: Status (large)
    c.setFillColor(status_text_color)
    c.setFont('Helvetica-Bold', 11)
    c.drawString(10, hdr_y_pt + hdr_h_pt - 14,
                 f"Page {page_num}  |  [{status_text}]  |  U: {u}  ({u_status})")

    # Line 2: Name + DNI
    c.setFillColor(GRAY_DARK)
    c.setFont('Helvetica', 9)
    c.drawString(10, hdr_y_pt + hdr_h_pt - 28,
                 f"Name: {expected_name or '(no name match)'}  |  DNI: {dni}")

    # Line 3: Answers
    c.setFont('Helvetica', 8)
    c.drawString(10, hdr_y_pt + hdr_h_pt - 42,
                 f"Answers: {n_answered}")

    # Line 4: Quality metrics
    c.setFillColor(GRAY)
    c.setFont('Helvetica', 7)
    c.drawString(10, hdr_y_pt + hdr_h_pt - 54,
                 f"Quality: {quality:.0%} ({n_boxes}/7 boxes)  |  "
                 f"Digit threshold: {digit_thr}  |  Answer threshold: {ans_thr}")

    # Manual-edit badge (top-right of header, if this page was hand-corrected)
    if r.get('_manual_edit'):
        badge_text = "[MANUALLY CORRECTED]"
        badge_font_size = 8
        c.setFont('Helvetica-Bold', badge_font_size)
        tw = c.stringWidth(badge_text, 'Helvetica-Bold', badge_font_size)
        bx = PAGE_W_PT - tw - 16
        by = hdr_y_pt + hdr_h_pt - 16
        c.setFillColor(ORANGE)
        c.rect(bx - 4, by - 3, tw + 8, badge_font_size + 6, fill=1, stroke=0)
        c.setFillColor(WHITE)
        c.drawString(bx, by, badge_text)

        legend_text = "Purple: O = added by reviewer | X = removed by reviewer | pill = field added/corrected by reviewer"
        legend_font_size = 6
        c.setFont('Helvetica', legend_font_size)
        c.setFillColor(PURPLE)
        c.drawString(10, hdr_y_pt + hdr_h_pt - 54 - 9, legend_text)

    # ========================================================
    # 4. Draw vector overlays on the IMAGE area
    # ========================================================
    bubble_hw_pt = px_w(BUBBLE_HALF_WIDTH)
    bubble_hh_pt = px_h(BUBBLE_HALF_HEIGHT)
    
    def draw_bubble_rect(cx_px, cy_px, color, line_w=0.8):
        """Draw a rectangle around a bubble at (cx_px, cy_px) in image coords."""
        x_pt, y_pt = px_to_pt(cx_px - BUBBLE_HALF_WIDTH, cy_px - BUBBLE_HALF_HEIGHT)
        c.setStrokeColor(color)
        c.setLineWidth(line_w)
        c.rect(x_pt, y_pt - 2 * bubble_hh_pt, 2 * bubble_hw_pt, 2 * bubble_hh_pt,
               stroke=1, fill=0)

    def draw_bubble_circle(cx_px, cy_px, color, line_w=0.9):
        """Circle a bubble: marks a reviewer added by hand (scanner saw no fill)."""
        x_pt, y_pt = px_to_pt(cx_px, cy_px)
        c.setStrokeColor(color)
        c.setLineWidth(line_w)
        c.ellipse(x_pt - bubble_hw_pt, y_pt - bubble_hh_pt,
                  x_pt + bubble_hw_pt, y_pt + bubble_hh_pt, stroke=1, fill=0)

    def draw_bubble_cross(cx_px, cy_px, color, line_w=0.9):
        """Cross out a bubble: a mark the scanner saw that a reviewer removed."""
        x_pt, y_pt = px_to_pt(cx_px, cy_px)
        c.setStrokeColor(color)
        c.setLineWidth(line_w)
        c.line(x_pt - bubble_hw_pt, y_pt - bubble_hh_pt, x_pt + bubble_hw_pt, y_pt + bubble_hh_pt)
        c.line(x_pt - bubble_hw_pt, y_pt + bubble_hh_pt, x_pt + bubble_hw_pt, y_pt - bubble_hh_pt)

    def draw_expected_slash(cx_px, cy_px, color, line_w=0.9):
        """Diagonal '/' pen-stroke over a bubble the answer key marks correct.

        Mirrors the "Show expected answers" overlay in the review GUI
        (gui/review_screen.py's _ExpectedAnswersLabel) so the exported page
        reads the same way.
        """
        x0, y0 = px_to_pt(cx_px - BUBBLE_HALF_WIDTH, cy_px + BUBBLE_HALF_HEIGHT)
        x1, y1 = px_to_pt(cx_px + BUBBLE_HALF_WIDTH, cy_px - BUBBLE_HALF_HEIGHT)
        c.setStrokeColor(color)
        c.setLineWidth(line_w)
        c.line(x0, y0, x1, y1)

    # --- Annotate ID field bubbles ---
    id_rows = r.get('_id_rows')
    median_sp = r.get('_median_sp', 50)
    x_offset = r.get('x_offset', 77)
    mask_arr = r.get('_mask')
    
    if id_rows is not None and mask_arr is not None:
        for cols in [REF_DNI_X, REF_CENTRE_X, REF_ASSIGN_X,
                     REF_PARCIAL_X, REF_PERMUT_X, REF_GRUP_X, REF_ID_X]:
            for ci, col in enumerate(cols):
                cx = col + x_offset
                d, _ = read_digit_marker_anchored(mask_arr, cx, id_rows,
                                                   median_sp, mask_arr.shape[1])
                if d is not None:
                    # Find the actual y of this bubble (using the same logic
                    # that decided it was detected)
                    y_margin = int(median_sp * 0.8)
                    y_top = int(id_rows[0]) - y_margin
                    y_bot = int(id_rows[-1]) + y_margin
                    x1 = max(0, cx - BUBBLE_HALF_WIDTH)
                    x2 = min(mask_arr.shape[1], cx + BUBBLE_HALF_WIDTH)
                    cs = mask_arr[y_top:y_bot, x1:x2]
                    if cs.size > 0:
                        pr = uniform_filter1d(cs.mean(axis=1) / 255.0, size=5)
                        pks, props = find_peaks(pr, height=0.03,
                                                 distance=int(median_sp * 0.4),
                                                 prominence=0.015)
                        if len(pks) > 0:
                            bp = pks[np.argmax(props['peak_heights'])]
                            py = bp + y_top
                            draw_bubble_rect(cx, py, GREEN, line_w=0.7)
                            # Digit label
                            x_pt, y_pt = px_to_pt(cx, py)
                            c.setFillColor(GREEN)
                            c.setFont('Helvetica-Bold', 5)
                            c.drawCentredString(x_pt, y_pt - 1, str(d))
    
    # --- Annotate answer bubbles ---
    answers = r.get('answers', {})
    for qn, adata in answers.items():
        col_x = adata.get('col_x', [])
        ay = adata.get('ans_y', 0)
        cy = adata.get('can_y', 0)
        af = adata.get('ans_fills', [])
        cf = adata.get('can_fills', [])
        current_marks = adata.get('marks', set())

        # Recompute what the scanner originally detected straight from the
        # saved pixel fill ratios (those never change after a manual edit),
        # so we can tell "OCR agrees" apart from "reviewer added/removed
        # this mark by hand" -- the fill ratios alone can't show that.
        auto_marks = set()
        for oi in range(len(col_x)):
            ans_marked = oi < len(af) and af[oi] > FILL_THRESHOLD_ANS
            can_marked = oi < len(cf) and cf[oi] > FILL_THRESHOLD_ANS
            if ans_marked and not can_marked:
                auto_marks.add(OPTION_LABELS[oi])

        added = current_marks - auto_marks
        removed = auto_marks - current_marks

        # Cancel-row marks (student crossed out a bubble): shown as-is, not
        # affected by manual answer edits.
        for oi in range(len(col_x)):
            if oi < len(cf) and cf[oi] > FILL_THRESHOLD_ANS:
                draw_bubble_rect(col_x[oi], cy, BLUE, line_w=0.6)
                x_pt, y_pt = px_to_pt(col_x[oi], cy)
                c.setFillColor(BLUE)
                c.setFont('Helvetica-Bold', 4)
                c.drawCentredString(x_pt, y_pt - 1, 'X')

        if not current_marks and not removed:
            continue

        score_info = adata.get('score_info', {})
        if score_info.get('all_correct'):
            color = GREEN
        elif score_info.get('partial'):
            color = YELLOW
        else:
            color = RED

        for oi in range(len(col_x)):
            opt = OPTION_LABELS[oi]
            if opt in added:
                # Reviewer added this mark; the scanner saw no fill here.
                draw_bubble_circle(col_x[oi], ay, PURPLE, line_w=0.9)
                x_pt, y_pt = px_to_pt(col_x[oi], ay)
                c.setFillColor(PURPLE)
                c.setFont('Helvetica-Bold', 5)
                c.drawCentredString(x_pt, y_pt - 1, opt)
            elif opt in current_marks:
                draw_bubble_rect(col_x[oi], ay, color, line_w=0.7)
                x_pt, y_pt = px_to_pt(col_x[oi], ay)
                c.setFillColor(color)
                c.setFont('Helvetica-Bold', 5)
                c.drawCentredString(x_pt, y_pt - 1, opt)
            elif opt in removed:
                # Scanner detected this mark; reviewer cancelled it by hand.
                draw_bubble_cross(col_x[oi], ay, PURPLE, line_w=0.9)

    # --- Expected-answer overlay (opt-in: only when the caller passes the
    # answer key, e.g. the single-student review export) -- drawn from the
    # key directly rather than inside the loop above, since a question with
    # no student marks would otherwise be skipped entirely by the
    # `if not current_marks and not removed: continue` short-circuit, and a
    # blank answer to a question the key marks correct is exactly the case
    # a reviewer most needs to see. ---
    if correct_answers_by_perm:
        permut_val = r.get('permut')
        correct_by_q = (correct_answers_by_perm.get(str(permut_val))
                         if permut_val is not None else None)
        if correct_by_q:
            for qn, correct_opts in correct_by_q.items():
                if not correct_opts:
                    continue
                adata = answers.get(qn)
                if not adata:
                    continue
                col_x = adata.get('col_x', [])
                ay = adata.get('ans_y')
                if ay is None:
                    continue
                for opt in correct_opts:
                    if opt not in OPTION_LABELS:
                        continue
                    oi = OPTION_LABELS.index(opt)
                    if oi < len(col_x):
                        draw_expected_slash(col_x[oi], ay, BLUE, line_w=0.9)

    # --- Annotate detected ID boxes ---
    boxes = r.get('_detected_boxes', {})
    permut_raw = r.get('permut')

    def _is_manually_edited(field):
        # '_pre_edit' is only set the first time a reviewer corrects this page
        # (see ReviewScreen._apply_correction); its absence means the value
        # shown is still the untouched OCR reading. It also backs the
        # "Revert to original" button.
        pre_edit = r.get('_pre_edit')
        if not pre_edit:
            return False
        return str(pre_edit.get(field) or '') != str(r.get(field) or '')

    def _overlay_color(field):
        # PURPLE takes priority: a reviewer's manual correction is the most
        # authoritative value regardless of how GRUP was originally derived.
        if _is_manually_edited(field):
            return PURPLE
        if field == 'grup':
            if str(r.get('grup_check', '')).startswith('MISMATCH'):
                return RED  # scanned GRUP disagrees with the roster -- needs a human look
            if r.get('grup_source') == 'roster':
                return TEAL  # bubble unreadable; backfilled from the roster
        return GREEN

    value_overlays = {
        'PARCIAL': (r.get('parcial', '') or '', _overlay_color('parcial')),
        'PERMUT': ('' if permut_raw is None else str(permut_raw), _overlay_color('permut')),
        'GRUP': (r.get('grup', '') or '', _overlay_color('grup')),
    }
    
    for name, box in boxes.items():
        bx, by, bw_px, bh_px = box['x'], box['y'], box['w'], box['h']
        # Box outline
        x1_pt, y1_pt = px_to_pt(bx, by)
        c.setStrokeColor(ORANGE)
        c.setLineWidth(0.5)
        c.rect(x1_pt, y1_pt - px_h(bh_px), px_w(bw_px), px_h(bh_px),
               stroke=1, fill=0)
        
        # Field name label (top-left of box)
        x_pt, y_pt = px_to_pt(bx + 2, by - 2)
        c.setFillColor(ORANGE)
        c.setFont('Helvetica-Bold', 4)
        c.drawString(x_pt, y_pt, name)
        
        # Value overlay for PARCIAL/PERMUT/GRUP
        if name in value_overlays:
            value, pill_color = value_overlays[name]
            if value and value != '' and value != 'None':
                # Pill above the box
                val_str = str(value)
                pill_font_size = 8.5  # ~20% bigger than the previous 7pt
                c.setFont('Helvetica-Bold', pill_font_size)
                tw = c.stringWidth(val_str, 'Helvetica-Bold', pill_font_size)
                cx_pt, cy_pt = px_to_pt(bx + bw_px // 2, by - 14)
                # Background pill
                pill_pad_x = 3
                pill_pad_y = 2
                c.setFillColor(PILL_BG)
                c.setStrokeColor(pill_color)
                c.setLineWidth(0.5)
                c.rect(cx_pt - tw/2 - pill_pad_x,
                       cy_pt - pill_pad_y,
                       tw + 2*pill_pad_x,
                       pill_font_size + 2*pill_pad_y,
                       stroke=1, fill=1)
                # Text
                c.setFillColor(pill_color)
                c.drawCentredString(cx_pt, cy_pt + 1, val_str)
    
    # ========================================================
    # 5. FOOTER with answer summary
    # ========================================================
    # The colour legend (opt-in, see correct_answers_by_perm/legend_extra_h
    # above -- the image was already shrunk to make room for this) gets an
    # extra strip of its own below the usual 36pt content, which is already
    # tight.
    footer_h_pt = 36 + legend_extra_h
    c.setFillColor(GRAY_LIGHT)
    c.rect(0, 0, PAGE_W_PT, footer_h_pt, fill=1, stroke=0)
    c.setStrokeColor(GRAY)
    c.setLineWidth(0.5)
    c.line(0, footer_h_pt, PAGE_W_PT, footer_h_pt)

    ans_strs = []
    for qn, adata in sorted(answers.items()):
        marks = adata.get('marks', set())
        if marks:
            ans_strs.append(f"Q{qn}={'+'.join(sorted(marks))}")
    ans_summary = "  ".join(ans_strs[:20])
    if len(ans_strs) > 20:
        ans_summary += f"  ... (+{len(ans_strs)-20} more)"

    c.setFillColor(GRAY_DARK)
    c.setFont('Helvetica-Bold', 7)
    c.drawString(10, 22 + legend_extra_h, "Answers detected:")
    c.setFillColor(GREEN)
    c.setFont('Helvetica', 7)
    c.drawString(10, 12 + legend_extra_h, ans_summary or "(none)")

    issues = r.get('validation_issues', [])
    if issues:
        c.setFillColor(RED)
        c.setFont('Helvetica', 6)
        c.drawString(10, 3 + legend_extra_h, "WARNINGS: " + "; ".join(issues)[:150])

    if show_legend:
        # Mirrors the colour legend shown under the preview panel in the
        # review GUI (gui/review_screen.py's _build_preview_panel), so the
        # exported page is self-explanatory without the app open.
        legend_items = [
            (GREEN, 'Correct'), (YELLOW, 'Partial'), (RED, 'Wrong'),
            (BLUE, 'Cancelled'), (PURPLE, 'Manual edit'),
        ]
        lx, ly = 10, 4
        swatch = 7
        c.setFont('Helvetica', 6)
        for color, label in legend_items:
            c.setFillColor(color)
            c.rect(lx, ly, swatch, swatch, fill=1, stroke=0)
            c.setFillColor(GRAY_DARK)
            c.drawString(lx + swatch + 2, ly + 1, label)
            lx += swatch + 2 + c.stringWidth(label, 'Helvetica', 6) + 10
        c.setStrokeColor(BLUE)
        c.setLineWidth(1.1)
        c.line(lx, ly, lx + swatch, ly + swatch)
        c.setFillColor(GRAY_DARK)
        c.drawString(lx + swatch + 2, ly + 1, "Expected (key)")

    c.showPage()


def write_annotated_pdf(all_results, output_path, students_df=None):
    """Generate multi-page annotated PDF with VECTOR overlays for review.

    Uses reportlab to compose each page:
      - Background: the scanned page (rasterized at moderate DPI to save space)
      - Overlay: vector graphics (lines, rectangles, text) that stay sharp at any zoom

    This is much better than burning annotations into the image:
      - Text is searchable and selectable in the PDF viewer
      - Annotations stay sharp at high zoom levels
      - File size is smaller (text is bytes, not pixels)
    """
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4

    PAGE_W_PT, PAGE_H_PT = A4  # 595 x 842 pts
    student_lookup = _build_student_name_lookup(students_df)

    c = canvas.Canvas(output_path, pagesize=A4)
    for r in all_results:
        _draw_annotated_page(c, r, student_lookup, PAGE_W_PT, PAGE_H_PT)
    c.save()
    return output_path


def _replace_with_retry(tmp_path, final_path, attempts=5, base_delay=0.3):
    """os.replace() with backoff.

    On Windows, replacing a file fails with WinError 5 (access denied) while
    anything else holds it open for reading -- including our own review GUI,
    which renders preview pages via a poppler subprocess in the background
    (see gui/review_screen.py's _PreviewRenderWorker). That read can overlap
    a save for the same file; retrying rides out the brief window instead of
    surfacing a spurious "could not save" error.
    """
    last_err = None
    for attempt in range(attempts):
        try:
            os.replace(tmp_path, final_path)
            return
        except OSError as e:
            last_err = e
            time.sleep(base_delay * (2 ** attempt))
    raise last_err


def patch_annotated_pdf_page(pdf_path, result, students_df, page_index):
    """Redraw a single page's annotation (after a manual correction) and splice
    it into the existing annotated_review.pdf at page_index (0-based), leaving
    every other page untouched. Much cheaper than regenerating the whole PDF,
    which matters when reviewing a large exam page by page.
    """
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from pypdf import PdfReader, PdfWriter
    import io

    PAGE_W_PT, PAGE_H_PT = A4
    student_lookup = _build_student_name_lookup(students_df)

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    _draw_annotated_page(c, result, student_lookup, PAGE_W_PT, PAGE_H_PT)
    c.save()
    buf.seek(0)
    new_page = PdfReader(buf).pages[0]

    reader = PdfReader(pdf_path)
    writer = PdfWriter()
    for i, page in enumerate(reader.pages):
        writer.add_page(new_page if i == page_index else page)

    tmp_path = pdf_path + '.tmp'
    with open(tmp_path, 'wb') as f:
        writer.write(f)
    _replace_with_retry(tmp_path, pdf_path)


def save_review_cache(all_results, students_df, correct_answers_by_perm,
                       num_questions, num_options, excel_path, pdf_path, cache_path,
                       exam_pdf_path=None, dpi=None):
    """Persist all OMR results to a pickle file for later review-session restores.

    What is persisted: all result dicts (with all fields), the student DataFrame,
    the answer key dict, question/option counts, and the output file paths.

    Why images are re-encoded:
    The '_corrected' numpy array and '_mask' array are stored as JPEG/PNG bytes
    rather than raw float32/uint8 arrays.  Raw arrays would be serialised by
    pickle as full numpy buffers (~3-5 MB per page), making a 100-page cache
    file 300-500 MB.  JPEG at quality=85 reduces each page to ~100-200 KB with
    no visible difference for review purposes; PNG (lossless) is used for the
    binary mask to preserve exact 0/255 values.

    excel_path and pdf_path are stored as basenames only (not full paths) so the
    cache survives the output folder being moved or renamed.  exam_pdf_path is
    stored as a full path since it lives outside the output folder.

    exam_pdf_path and dpi are optional for backward-compatibility with cache
    files written by older versions that did not include these fields.
    """
    import pickle

    cached_results = []
    for r in all_results:
        cr = {k: v for k, v in r.items() if k not in ('_corrected', '_mask')}
        corr = r.get('_corrected')
        if corr is not None:
            ok, buf = cv2.imencode('.jpg', corr, [cv2.IMWRITE_JPEG_QUALITY, 85])
            cr['_corrected_jpg'] = buf.tobytes() if ok else None
        mask = r.get('_mask')
        if mask is not None:
            ok, buf = cv2.imencode('.png', mask)
            cr['_mask_png'] = buf.tobytes() if ok else None
        cached_results.append(cr)

    payload = {
        'version': 1,
        'all_results': cached_results,
        'students_df': students_df,
        'correct_answers_by_perm': correct_answers_by_perm,
        'num_questions': num_questions,
        'num_options': num_options,
        # Stored as filenames, not full paths: lets the cache survive the
        # output folder being moved or renamed (e.g. inside synced Drive
        # folders), as long as the three files stay together.
        'excel_filename': os.path.basename(excel_path),
        'pdf_filename': os.path.basename(pdf_path),
        # Full path, unlike excel/pdf: it lives outside the output folder
        # (wherever the user's original scan happens to be), so there's no
        # "stays next to the cache" guarantee to lean on.
        'exam_pdf_path': exam_pdf_path,
        'dpi': dpi,
    }
    with open(cache_path, 'wb') as f:
        pickle.dump(payload, f, protocol=pickle.HIGHEST_PROTOCOL)
    return cache_path


def load_review_cache(cache_path):
    """Restore a previously saved OMR session from a pickle cache file.

    Inverse of save_review_cache(): decodes the per-page JPEG/PNG bytes back
    into numpy arrays ('_corrected' and '_mask') and reconstructs the full
    run_state dict expected by the Review window.

    Output file paths are resolved relative to the cache file's own directory
    (since they were stored as basenames), so the whole output folder can be
    moved without breaking path lookups.  The 'cache_path' key in the returned
    dict is always the absolute path to the cache file itself, ensuring that
    manual edits made during the review session get written back to the same
    file rather than creating a new one.
    """
    import pickle

    with open(cache_path, 'rb') as f:
        payload = pickle.load(f)

    all_results = []
    for cr in payload['all_results']:
        r = dict(cr)
        jpg = r.pop('_corrected_jpg', None)
        if jpg is not None:
            r['_corrected'] = cv2.imdecode(np.frombuffer(jpg, dtype=np.uint8), cv2.IMREAD_COLOR)
        png = r.pop('_mask_png', None)
        if png is not None:
            r['_mask'] = cv2.imdecode(np.frombuffer(png, dtype=np.uint8), cv2.IMREAD_UNCHANGED)
        all_results.append(r)

    cache_dir = os.path.dirname(os.path.abspath(cache_path))
    return {
        'all_results': all_results,
        'students_df': payload['students_df'],
        'correct_answers_by_perm': payload['correct_answers_by_perm'],
        'num_questions': payload['num_questions'],
        'num_options': payload['num_options'],
        # Without this, edits made after reopening a previous session's
        # results would patch results.xlsx/annotated_review.pdf but never
        # update review_cache.pkl itself -- so the *next* reopen would lose
        # every correction made in this session.
        'cache_path': os.path.abspath(cache_path),
        'excel_path': os.path.join(cache_dir, payload['excel_filename']),
        'pdf_path': os.path.join(cache_dir, payload['pdf_filename']),
        'exam_pdf': payload.get('exam_pdf_path'),
        'dpi': payload.get('dpi'),
    }


def _read_answer_key_dataframe(answers_path):
    """Load the raw Perm/QuestionNum/option-columns table from CSV or Excel.

    Shared by load_correct_answers() and validate_answer_key() so validation
    always sees exactly the rows the scoring pipeline will use -- including
    the same row order, which the mislabeled-Perm heuristic in
    validate_answer_key() depends on.
    """
    if answers_path.lower().endswith('.csv'):
        df = pd.read_csv(answers_path)
    else:
        df = pd.read_excel(answers_path)

    df.columns = [str(c).strip() for c in df.columns]

    if 'Perm' not in df.columns or 'QuestionNum' not in df.columns:
        raise ValueError(
            "Answers file must have 'Perm' and 'QuestionNum' columns "
            "(one row per question per permutation, with option columns "
            "A, B, C, ... holding 1/0). Found columns: " + ', '.join(df.columns)
        )
    return df


def validate_answer_key(answers_path, expected_num_questions=None, expected_num_options=None):
    """Check an answer-key file for data-entry mistakes that would otherwise
    silently corrupt grading.

    load_correct_answers() and the scoring code are both permissive by
    design: a (Perm, QuestionNum) pair that's missing is just skipped (no
    error, the question is excluded from that permutation's max score) and a
    duplicate pair silently lets the last row win. That's convenient for
    normal use but means a copy/paste slip in the source spreadsheet -- e.g.
    a row's Perm value bumped one row too early -- produces wrong or
    incomplete grades with zero indication anything is wrong.

    This function re-parses the same file and flags four problem classes:
      - BLANK_PERM:   a row's Perm cell is blank/NaN. Flagged on its own
                      rather than left to become the literal string "nan"
                      and silently masquerade as a real permutation (see
                      the blank-Perm handling below for what that caused).
      - DUPLICATE:    the same (Perm, QuestionNum) pair appears more than once.
      - MISSING:      a permutation is missing a QuestionNum that every other
                      permutation in the file has (permutations are assumed to
                      all cover the same question numbers).
      - MISSING_OPTION: the file has fewer option columns (A, B, C, ...) than
                      expected_num_options says the exam actually has. This
                      one is easy to hit and silently devastating: option
                      columns are read straight from whatever's in the file
                      (load_correct_answers() has no idea how many options
                      the exam is *supposed* to have), so a stale/wrong
                      template -- e.g. a 4-option CSV reused for a 6-option
                      exam -- costs every student credit for a genuinely
                      correct E or F mark with zero warning: that option is
                      just never in any question's correct set, so marking
                      it scores as if it were wrong. Confirmed by testing.

    For missing questions, it also looks for a plausible cause: a row
    elsewhere whose Perm changed but whose QuestionNum kept incrementing
    instead of resetting to 1 (the signature of a Perm value copy/pasted one
    row too early in a file ordered by Perm-then-QuestionNum). When found,
    the issue's 'suggested_fix' names the exact row and the Perm value it
    was probably meant to have.

    Args:
        answers_path: path to the CSV/Excel answers file
        expected_num_questions: if given, the question set every permutation
            is checked against is 1..expected_num_questions instead of the
            union of QuestionNum values actually seen in the file (catches a
            permutation that's missing its last question(s) entirely, which
            a union-only check would miss)
        expected_num_options: if given, checks that option columns A, B, C...
            up to this many exist in the file (catches a stale/wrong-sized
            template, e.g. a 4-option file used for a 6-option exam)

    Returns:
        list of issue dicts, each with:
            'severity'      -- always 'error' (every class here silently
                                corrupts grading; there is no lesser rank yet)
            'perm'          -- affected permutation key (str), or None for a
                                file-wide issue (e.g. MISSING_OPTION)
            'question'      -- affected question number (int), or None
            'message'       -- human-readable description
            'rows'          -- list of 1-based file row numbers involved
                                (header is row 1)
            'suggested_fix' -- None, or a dict:
                {'row': int, 'new_perm': str, 'description': str}
    """
    df = _read_answer_key_dataframe(answers_path)
    df = df.copy()

    issues = []

    # --- MISSING_OPTION: fewer option columns than the exam is configured
    # for. A file-wide check (not per-permutation), done before anything
    # else since it doesn't depend on Perm/QuestionNum at all.
    if expected_num_options:
        option_cols = {c for c in df.columns if c not in ('Perm', 'QuestionNum')}
        expected_letters = OPTION_LABELS[:expected_num_options]
        missing_letters = [l for l in expected_letters if l not in option_cols]
        if missing_letters:
            issues.append({
                'severity': 'error',
                'perm': None,
                'question': None,
                'message': (
                    f"The answer key only has columns for options "
                    f"{', '.join(sorted(option_cols & set(expected_letters))) or '(none)'}, "
                    f"but this exam is configured for {expected_num_options} "
                    f"options ({', '.join(expected_letters)}). Missing: "
                    f"{', '.join(missing_letters)} -- a student marking one "
                    f"of those bubbles will be silently scored as if it were "
                    f"always wrong, since it can never appear in any "
                    f"question's correct-answer set."
                ),
                'rows': [],
                'suggested_fix': None,
            })

    # --- Blank Perm: flag distinctly and drop from every check below,
    # rather than letting _normalize_perm_value() turn a blank/NaN cell
    # into the literal string "nan" and have it flow through as if it
    # were a real permutation. Confirmed by testing: left unflagged, this
    # produces a confusing "Perm nan is missing Question N" message from
    # the missing-question check below, AND a spurious "Perm nan" sheet in
    # the Excel output (write_excel creates one sheet per key in the
    # loaded answers dict).
    blank_perm_mask = df['Perm'].isna() | (df['Perm'].astype(str).str.strip() == '')
    for idx in df.index[blank_perm_mask]:
        q = df.loc[idx, 'QuestionNum']
        q_display = int(q) if pd.notna(pd.to_numeric(q, errors='coerce')) else str(q)
        issues.append({
            'severity': 'error',
            'perm': None,
            'question': None,
            'message': (f"Row {idx + 2} has a blank Perm value (QuestionNum "
                        f"{q_display}) -- every row must belong to a "
                        f"permutation. Fix or delete this row."),
            'rows': [idx + 2],
            'suggested_fix': None,
        })
    df = df[~blank_perm_mask].copy()

    df['Perm'] = df['Perm'].apply(_normalize_perm_value)
    df['QuestionNum'] = pd.to_numeric(df['QuestionNum'], errors='coerce')

    # --- Duplicates: same (Perm, QuestionNum) appears more than once ---
    dup_mask = df.duplicated(subset=['Perm', 'QuestionNum'], keep=False) & df['QuestionNum'].notna()
    for (perm, q), group in df[dup_mask].groupby(['Perm', 'QuestionNum']):
        rows = [int(idx) + 2 for idx in group.index]
        issues.append({
            'severity': 'error',
            'perm': perm,
            'question': int(q),
            'message': (f"Perm {perm}, Question {int(q)} appears {len(rows)} times "
                        f"(rows {', '.join(map(str, rows))}). Only the last one is used "
                        f"for scoring -- the rest are silently discarded."),
            'rows': rows,
            'suggested_fix': None,
        })

    # --- Heuristic: a row whose Perm differs from the previous row but whose
    # QuestionNum keeps incrementing (instead of resetting to 1, as it
    # normally would at the start of a new permutation's block) looks like a
    # Perm value that was copy/pasted one row too early. ---
    mislabel_suspects = {}  # (perm, q) -> (suggested_perm, row_number)
    prev_perm, prev_q = None, None
    for idx, row in df.iterrows():
        perm, q = row['Perm'], row['QuestionNum']
        if (prev_perm is not None and perm != prev_perm
                and pd.notna(q) and pd.notna(prev_q) and q == prev_q + 1):
            mislabel_suspects[(perm, q)] = (prev_perm, idx + 2)
        prev_perm, prev_q = perm, q

    # --- Missing questions: every permutation should cover the same set of
    # question numbers. ---
    if expected_num_questions:
        expected_qnums = list(range(1, expected_num_questions + 1))
    else:
        expected_qnums = sorted(int(q) for q in df['QuestionNum'].dropna().unique())

    for perm, group in df.groupby('Perm'):
        have = set(group['QuestionNum'].dropna().astype(int))
        for q in expected_qnums:
            if q in have:
                continue
            suggested_fix = None
            for (s_perm, s_q), (suggested_perm, row_no) in mislabel_suspects.items():
                if int(s_q) == q and suggested_perm == perm:
                    suggested_fix = {
                        'row': row_no,
                        'new_perm': suggested_perm,
                        'description': (
                            f"Row {row_no} is labeled Perm={s_perm}, QuestionNum={q}, "
                            f"but it appears immediately after Perm {suggested_perm}'s "
                            f"Question {q - 1} -- it looks like it should be "
                            f"Perm={suggested_perm} instead."
                        ),
                    }
                    break
            issues.append({
                'severity': 'error',
                'perm': perm,
                'question': q,
                'message': (f"Perm {perm} is missing Question {q}. It will be silently "
                            f"skipped when scoring this permutation (no error -- just a "
                            f"smaller max possible grade for every student on this Perm)."),
                'rows': [],
                'suggested_fix': suggested_fix,
            })

    # 'question' is None for a blank-Perm issue (there's no question it's
    # scoped to); sort those last within their group instead of comparing
    # None to an int, which raises TypeError as soon as two such issues
    # tie on str(perm) (confirmed: 2+ blank-Perm rows crashed this sort).
    issues.sort(key=lambda i: (str(i['perm']), i['question'] if i['question'] is not None else -1))
    return issues


def apply_answer_key_row_fix(answers_path, row_number, new_perm):
    """Apply a single suggested_fix from validate_answer_key(): relabel one
    row's Perm value in place.

    A backup of the original file is written alongside it (same path with a
    '.bak' suffix, overwritten on repeated calls) before any modification,
    since this edits the user's original input file.

    Args:
        answers_path: path to the CSV/Excel answers file
        row_number: 1-based file row to edit (header is row 1), as returned
            in a suggested_fix dict
        new_perm: the Perm value to write into that row
    """
    import shutil
    backup_path = answers_path + '.bak'
    shutil.copy2(answers_path, backup_path)

    df = _read_answer_key_dataframe(answers_path)
    data_idx = row_number - 2
    if data_idx < 0 or data_idx >= len(df):
        raise ValueError(f"Row {row_number} is out of range for {answers_path}")

    perm_col = 'Perm' if 'Perm' in df.columns else [c for c in df.columns if c.strip() == 'Perm'][0]
    # The Perm column is often read as int64 (e.g. bare "1"/"2" values); a
    # plain string assignment into an int64 column raises in modern pandas.
    # Match the existing dtype when the new value parses cleanly as an int,
    # otherwise widen the column to object so non-numeric Perm labels work.
    if pd.api.types.is_integer_dtype(df[perm_col].dtype):
        try:
            new_val = int(new_perm)
        except (TypeError, ValueError):
            df[perm_col] = df[perm_col].astype(object)
            new_val = new_perm
    else:
        new_val = new_perm
    df.loc[data_idx, perm_col] = new_val

    if answers_path.lower().endswith('.csv'):
        df.to_csv(answers_path, index=False)
    else:
        df.to_excel(answers_path, index=False)


def load_correct_answers(answers_path):
    """Load correct answers for all exam permutations from an Excel/CSV file.

    The file uses a Perm x QuestionNum matrix format: each row describes one
    question of one permutation, and each option column (A, B, C, ...) holds
    1 (correct) or 0 (incorrect) for that question.  Multiple correct options
    per question are supported for multi-answer scoring.

    Required column names (whitespace around values is stripped automatically):
        Perm        -- permutation identifier (e.g. 0, 1, 2)
        QuestionNum -- 1-based question number
        A, B, C, D, [E, ...] -- 1/0 correctness per option

    Any number of permutations can be present; they are grouped by the 'Perm'
    column value.  The permutation key is stored as a string so it matches the
    PERMUT bubble readout (also string) in write_excel's routing step.

    Silently tolerates the missing/duplicate-question data errors that
    validate_answer_key() flags (a missing question is just absent from the
    returned dict; a duplicate keeps the last row) -- callers that care
    should run validate_answer_key() first and surface issues to the user.
    A blank Perm cell is also tolerated the same way: the row is skipped
    rather than becoming a bogus permutation labeled "nan" (which would
    otherwise show up as a spurious extra sheet in write_excel's output).

    Returns:
        dict mapping perm_key (str) -> {question_num (int): set_of_correct_letters}
    """
    df = _read_answer_key_dataframe(answers_path)

    option_cols = [c for c in df.columns if c not in ('Perm', 'QuestionNum')]

    correct_by_perm = {}
    for _, row in df.iterrows():
        if pd.isna(row['Perm']) or str(row['Perm']).strip() == '':
            continue
        perm = _normalize_perm_value(row['Perm'])
        q = int(row['QuestionNum'])
        correct_set = set()
        for opt_col in option_cols:
            try:
                val_num = float(str(row[opt_col]).strip())
            except (ValueError, TypeError):
                val_num = 0.0
            if not pd.isna(val_num) and val_num != 0:
                correct_set.add(opt_col)
        correct_by_perm.setdefault(perm, {})[q] = correct_set

    return correct_by_perm


def load_students(students_path):
    """Load and normalize a student list from Excel or CSV.

    Four supported formats:

    Format 1 -- UPF official .xls export (semicolon-separated or Excel):
        Row 1: course name (single cell, not a column header)
        Row 2: IDUSUARI ; NIA ; NIP ; COGNOM1 ; COGNOM2 ; NOM
        Row 3+: data
        The IDUSUARI column contains the U-number (e.g. "U123456").
        NIA, NIP, and any other extra columns are silently dropped.

    Format 2 -- Simple CSV or Excel with conventional column names:
        Nom | Cognom1 | Cognom2 | U_number
        Column name matching is accent- and case-insensitive.

    Format 3 -- Moodle "participants" export (e.g. courseid_NNNNN_participants.csv):
        Nom | Cognoms | "Número ID" | Grups
        Cognoms merges both surnames into one field (split on the first
        space: everything before is Cognom1, everything after is Cognom2 --
        imperfect for 3+-word compound surnames, but there's no reliable way
        to split those without a name database, and it only affects display).
        "Número ID" holds the U-number, typically lowercase-prefixed
        ("u123456"); the 'U' is stripped downstream like any other format.
        Grups encodes theory group + seminar subgroup as "<g><nn>-<s>" (e.g.
        "201-7"): the theory group is always the single leading digit (1 or
        2 in every export seen so far), so it's parsed out into a
        'TheoryGroup' column. There's no field for the seminar subgroup on
        the exam bubble sheet, so that part is discarded. TheoryGroup is
        used by backfill_and_validate_groups() to cross-check/backfill the
        scanned GRUP field.

    Format 4 -- UPF official .xls export with EMAIL/PRACTICA (e.g.
    "llistatGGiA (N).xls"): same two-row layout as Format 1
    (course title, then IDUSUARI;NIA;COGNOM1;COGNOM2;NOM;EMAIL;PRACTICA), plus:
        EMAIL     -- kept as-is in an 'Email' column (not used for grading,
                     but handy for emailing a student their scanned exam)
        PRACTICA  -- a 3-digit code (e.g. "102"); its leading digit is the
                     theory group, parsed into 'TheoryGroup' the same way as
                     Format 3's "Grups" column

    The function auto-detects whether the first data row is actually a header
    (UPF format) by checking if any of the recognized column names appear in
    that row.  CSV separator (comma vs semicolon) and encoding (UTF-8, Latin-1,
    cp1252) are also sniffed automatically.

    Returns:
        DataFrame with columns: Nom, Cognom1, Cognom2, U_number, plus Email
        if the source had one, and TheoryGroup if the source had a
        Grups/Grup/Practica column with parseable data.
        Rows with empty/null U_number are dropped.
    """
    # Normalize column names: lowercase + strip accents. Defined before the
    # CSV-reading block (not just before its later main use) because the
    # headerless fallback below needs it to find the real header row.
    def normalize_col(c):
        s = str(c).strip().lower()
        # Remove accents
        for a, b in [('à','a'),('á','a'),('è','e'),('é','e'),('í','i'),
                     ('ó','o'),('ò','o'),('ú','u'),('ü','u'),('ñ','n'),('ç','c')]:
            s = s.replace(a, b).replace(a.upper(), b.upper())
        return s

    # Pick correct reader based on extension
    ext = students_path.lower().split('.')[-1]
    if ext == 'csv':
        # Auto-detect separator (comma vs semicolon) and encoding (UTF-8 vs Latin-1)
        # Try common combinations until one works without errors
        attempts = [
            {'sep': ',', 'encoding': 'utf-8'},
            {'sep': ';', 'encoding': 'utf-8'},
            {'sep': ',', 'encoding': 'latin-1'},
            {'sep': ';', 'encoding': 'latin-1'},
            {'sep': ',', 'encoding': 'cp1252'},
            {'sep': ';', 'encoding': 'cp1252'},
        ]
        # First, sniff the separator from the file content
        try:
            with open(students_path, 'rb') as f:
                head = f.read(2000)
            # Decode for sniffing (try latin-1 since it can decode any byte)
            sample = head.decode('latin-1', errors='replace')
            # If semicolons are more common than commas in the first lines, prefer ;
            first_lines = sample.split('\n')[:3]
            joined = '\n'.join(first_lines)
            n_semi = joined.count(';')
            n_comma = joined.count(',')
            if n_semi > n_comma:
                # Reorder to try semicolon first
                attempts = sorted(attempts, key=lambda a: 0 if a['sep'] == ';' else 1)
        except Exception:
            pass
        
        df = None
        last_err = None
        for opts in attempts:
            try:
                df_try = pd.read_csv(students_path, **opts)
                # Must have at least 2 columns to be valid
                if df_try.shape[1] >= 2:
                    df = df_try
                    break
            except Exception as e:
                last_err = e

        if df is None:
            # A genuine "UPF official export" CSV (course-title-only line 1,
            # real headers on line 2 -- this function's own docstring
            # documents that variant as supported) can have ZERO separators
            # on that first line, so every attempt above -- which lets
            # pandas infer columns from line 1 -- sees exactly 1 column no
            # matter which separator is tried, and shape[1] >= 2 never
            # passes even when the separator guess was correct. Confirmed
            # by testing: this made that documented format unconditionally
            # fail to load.
            #
            # header=None does NOT fix this the way it might seem to:
            # pandas' C parser still infers the expected field count from
            # line 1 even with no declared header, and raises ParserError
            # on every subsequent (wider) line instead of padding -- also
            # confirmed by testing. So instead, read the raw lines directly
            # to find which one actually looks like the header, then have
            # pandas re-read starting from exactly that line via `skiprows`
            # -- this sidesteps the ragged-row problem entirely, since
            # pandas never sees the mismatched title line at all.
            for opts in attempts:
                try:
                    with open(students_path, 'r', encoding=opts['encoding'], errors='replace') as f:
                        raw_lines = [line for _, line in zip(range(5), f)]
                except Exception as e:
                    last_err = e
                    continue
                header_line_idx = None
                for i, line in enumerate(raw_lines):
                    fields = [normalize_col(v) for v in line.rstrip('\r\n').split(opts['sep'])]
                    if any(c in fields for c in ['idusuari', 'cognom1', 'nom']):
                        header_line_idx = i
                        break
                if header_line_idx is None:
                    continue
                try:
                    df_try = pd.read_csv(students_path, skiprows=header_line_idx, **opts)
                except Exception as e:
                    last_err = e
                    continue
                if df_try.shape[1] < 2:
                    continue
                df = df_try
                break

        if df is None:
            raise RuntimeError(f"Could not read CSV with any standard format. Last error: {last_err}")
    elif ext == 'xls':
        # Old Excel format requires xlrd engine
        try:
            df = pd.read_excel(students_path, engine='xlrd')
        except ImportError:
            raise RuntimeError(
                "Reading .xls files requires xlrd. Install with: pip install xlrd"
            )
    else:  # xlsx, xlsm
        df = pd.read_excel(students_path)

    # Detect if first row is a course title (single-cell header that spans columns)
    # The UPF export has the course code in row 0, real headers in row 1
    cols_normalized = [normalize_col(c) for c in df.columns]
    looks_like_data_header = any(
        c in cols_normalized for c in ['nom', 'cognom1', 'u_number', 'idusuari', 'nip', 'nia']
    )
    
    if not looks_like_data_header and len(df) > 0:
        # First row may be the real header
        first_row = [normalize_col(v) for v in df.iloc[0].tolist()]
        if any(c in first_row for c in ['idusuari', 'cognom1', 'nom']):
            # Use row 0 as header, drop it from data
            df.columns = df.iloc[0].tolist()
            df = df.iloc[1:].reset_index(drop=True)
            cols_normalized = [normalize_col(c) for c in df.columns]
    
    # Build mapping: original_col -> normalized_target
    col_mapping = {}
    for orig, norm in zip(df.columns, cols_normalized):
        if norm in ('nom',):
            col_mapping[orig] = 'Nom'
        elif norm in ('cognom1', 'cognom 1', 'primer cognom', 'apellido1', 'apellido 1'):
            col_mapping[orig] = 'Cognom1'
        elif norm in ('cognom2', 'cognom 2', 'segon cognom', 'apellido2', 'apellido 2'):
            col_mapping[orig] = 'Cognom2'
        elif norm in ('u_number', 'unumber', 'u-number', 'idusuari', 'id_usuari',
                      'numero id', 'numeroid'):
            col_mapping[orig] = 'U_number'
        elif norm == 'cognoms':
            col_mapping[orig] = 'Cognoms'
        elif norm in ('grups', 'grup', 'practica'):
            # Both UPF's "Grups" ("201-7") and "PRACTICA" ("102") formats
            # encode the theory group as their leading digit; the rest
            # (seminar subgroup / lab practice number) isn't used here.
            col_mapping[orig] = 'GrupRaw'
        elif norm in ('email', 'e-mail', 'correu', 'correu electronic', 'correu-e'):
            col_mapping[orig] = 'Email'

    df = df.rename(columns=col_mapping)

    # Format 3: single merged "Cognoms" column -> split into Cognom1/Cognom2
    if 'Cognoms' in df.columns and 'Cognom1' not in df.columns:
        parts = df['Cognoms'].astype(str).str.strip().str.split(n=1, expand=True)
        df['Cognom1'] = parts[0]
        df['Cognom2'] = parts[1] if parts.shape[1] > 1 else ''
        df['Cognom2'] = df['Cognom2'].fillna('')

    # Format 3/4: "Grups"/"PRACTICA" -> theory group is the leading digit
    if 'GrupRaw' in df.columns:
        leading_digit = df['GrupRaw'].astype(str).str.strip().str[:1]
        df['TheoryGroup'] = leading_digit.where(leading_digit.isin(['1', '2']))

    # Ensure all required columns exist
    for required in ['Nom', 'Cognom1', 'Cognom2', 'U_number']:
        if required not in df.columns:
            if required == 'Cognom2':
                df[required] = ''  # Optional
            else:
                raise ValueError(
                    f"Required column '{required}' not found in students file. "
                    f"Found columns: {list(df.columns)}"
                )

    # Keep only the columns we care about (drop NIA, NIP, Grups, etc.)
    keep_cols = ['Nom', 'Cognom1', 'Cognom2', 'U_number']
    if 'Email' in df.columns:
        keep_cols.append('Email')
    if 'TheoryGroup' in df.columns:
        keep_cols.append('TheoryGroup')
    df = df[keep_cols].copy()

    # A present-but-empty cell reads back as NaN, not '' -- left as-is it
    # renders as the literal text "nan" wherever the value is interpolated
    # into a display string (Excel, annotated PDF, email templates).
    # Cognom2 (no second surname) is the common case, but testing showed
    # Nom/Cognom1/Email are equally affected by any blank source cell.
    for col in ('Nom', 'Cognom1', 'Cognom2', 'Email'):
        if col in df.columns:
            df[col] = df[col].fillna('')

    # Clean: drop rows where U_number is empty/nan
    df['U_number'] = df['U_number'].astype(str).str.strip()
    df = df[df['U_number'].notna() & (df['U_number'] != '') & 
            (df['U_number'].str.lower() != 'nan')]
    df = df.reset_index(drop=True)
    
    return df


# =============================================
# CLI ENTRY POINT
# =============================================

def detect_pdf_dpi(pdf_path):
    """Auto-detect the embedded image resolution of a scanned PDF.

    Why this matters: process_page() rescales every page to REFERENCE_DPI
    (300) internally, but only AFTER convert_from_path() has already rendered
    the PDF at `source_dpi`.  If the PDF's embedded images are at 600 DPI and
    we render at 300 DPI, pdf2image downsamples by 2x and we lose resolution;
    if we render at 600 DPI when the images are 300 DPI, we get upsampled
    output that wastes memory and slows down processing.  Matching the render
    DPI to the PDF's actual image DPI avoids both problems.

    Uses poppler's `pdfimages -list` to read the x-ppi field from the first
    embedded image.  The median of all images is snapped to the nearest
    standard scanner DPI (150/300/600/1200) to be robust against minor
    metadata rounding.

    Returns 300 as a safe default if poppler is unavailable or the PDF has
    no embedded images (e.g. vector-only or password-protected PDFs).
    """
    try:
        import subprocess
        result = subprocess.run(['pdfimages', '-list', pdf_path],
                                capture_output=True, text=True, timeout=10)
        if result.returncode == 0:
            # Parse pdfimages output - look for x-ppi column
            lines = result.stdout.strip().split('\n')
            if len(lines) >= 3:  # Header + separator + at least one image
                # Find DPI values in the data rows
                dpis = []
                for line in lines[2:]:
                    parts = line.split()
                    # x-ppi is typically column 12-13 in pdfimages output
                    for part in parts:
                        try:
                            val = int(part)
                            if 100 <= val <= 1200:  # Reasonable DPI range
                                dpis.append(val)
                                break
                        except ValueError:
                            continue
                
                if dpis:
                    median_dpi = sorted(dpis)[len(dpis) // 2]
                    # Round to nearest common DPI value
                    if median_dpi < 200:
                        return 150
                    elif median_dpi < 400:
                        return 300
                    elif median_dpi < 800:
                        return 600
                    else:
                        return 1200
        
        return 300  # Default fallback
    except Exception:
        return 300


def main():
    parser = argparse.ArgumentParser(
        description='OMR Exam Corrector for UPF multi-answer tests',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python omr_correct.py exams.pdf students.csv answers.csv --questions 10
  python omr_correct.py exams.pdf students.csv answers.csv -q 20 -o results/

Answer file format (CSV/Excel) - one row per question per exam permutation,
with a mandatory 'Perm' column and one 1/0 column per option:
  Perm | QuestionNum | A | B | C | D
  0    | 1           | 0 | 1 | 0 | 0
  0    | 2           | 1 | 1 | 0 | 0
  1    | 1           | 0 | 1 | 0 | 0
  ...

Students file format (CSV/Excel) - either:
  Nom | Cognom1 | Cognom2 | U_number
  ...
or the official UPF export (IDUSUARI;NIA;NIP;COGNOM1;COGNOM2;NOM), detected
automatically.
"""
    )
    parser.add_argument('exam_pdf', help='PDF with scanned exams')
    parser.add_argument('students', help='Excel/CSV with student list')
    parser.add_argument('answers', help='Excel/CSV with correct answers')
    parser.add_argument('-q', '--questions', type=int, required=True,
                       help='Number of questions in the exam')
    parser.add_argument('-n', '--num-options', type=int, default=5,
                       help='Number of options per question (default: 5)')
    parser.add_argument('-o', '--output-dir', default='./output',
                       help='Output directory (default: ./output)')
    parser.add_argument('--dpi', type=int, default=0,
                       help='PDF rendering DPI (default: auto-detect, typically 300 or 600). '
                            'Use --dpi 300 or --dpi 600 to force.')
    parser.add_argument('-v', '--verbose', action='store_true',
                       help='Verbose output')
    parser.add_argument('--ignore-answer-key-warnings', action='store_true',
                       help='Proceed even if the answers file has missing/duplicate '
                            'QuestionNum rows for a permutation (grading would silently '
                            'be wrong or incomplete for that permutation).')

    args = parser.parse_args()
    
    # Validate num_options
    if args.num_options < 2 or args.num_options > len(OPTION_LABELS):
        print(f"ERROR: num-options must be between 2 and {len(OPTION_LABELS)} "
              f"(got {args.num_options})")
        sys.exit(1)
    
    # Validate paths
    if not os.path.exists(args.exam_pdf):
        print(f"ERROR: Exam PDF not found: {args.exam_pdf}")
        sys.exit(1)
    if not os.path.exists(args.students):
        print(f"ERROR: Students file not found: {args.students}")
        sys.exit(1)
    if not os.path.exists(args.answers):
        print(f"ERROR: Answers file not found: {args.answers}")
        sys.exit(1)
    
    os.makedirs(args.output_dir, exist_ok=True)
    
    # Load inputs
    print(f"Loading students from {args.students}...")
    students_df = load_students(args.students)
    print(f"  {len(students_df)} students loaded")
    
    print(f"Loading correct answers from {args.answers}...")
    issues = validate_answer_key(args.answers, expected_num_questions=args.questions,
                                  expected_num_options=args.num_options)
    if issues:
        print(f"\n  WARNING: {len(issues)} problem(s) found in the answers file:")
        for issue in issues:
            print(f"    - {issue['message']}")
            if issue['suggested_fix']:
                print(f"      Suggested fix: {issue['suggested_fix']['description']}")
        if not args.ignore_answer_key_warnings:
            print("\n  Refusing to proceed with a possibly-wrong answer key.")
            print("  Fix the file above, or re-run with --ignore-answer-key-warnings "
                  "to proceed anyway.")
            sys.exit(1)
        print("  --ignore-answer-key-warnings set: proceeding anyway.\n")

    correct_answers_by_perm = load_correct_answers(args.answers)
    perm_names = ', '.join(sorted(correct_answers_by_perm.keys()))
    print(f"  {len(correct_answers_by_perm)} permutation(s) loaded: {perm_names}")
    
    # Auto-detect DPI if not specified
    if args.dpi <= 0:
        print(f"Auto-detecting source DPI...")
        args.dpi = detect_pdf_dpi(args.exam_pdf)
        print(f"  Detected: {args.dpi} DPI")
    
    # Process PDF
    print(f"Rendering {args.exam_pdf} at {args.dpi} DPI...")
    pages = convert_from_path(args.exam_pdf, dpi=args.dpi)
    print(f"  {len(pages)} pages rendered")
    
    print(f"\nProcessing pages with perspective correction...")
    all_results = []
    for i, page in enumerate(pages):
        try:
            r, corr = process_page(page, i+1, args.questions, args.num_options,
                                    source_dpi=args.dpi)
            all_results.append(r)

            u = r.get('u_number', '?')
            dni = r.get('dni', '?')
            n_ans = sum(1 for a in r.get('answers', {}).values() if a.get('marks'))
            # ASCII, not a unicode checkmark: Windows consoles commonly default
            # to a cp1252-family codepage that can't encode U+2713/U+2717,
            # which crashed the whole run mid-batch the first time this
            # printed a failed page.
            status = 'OK' if r.get('status') == 'OK' else 'FAIL'
            print(f"  [{status}] P{i+1:3d}: DNI={dni:>10s} U={str(u):>10s} answers={n_ans}")
        except Exception as e:
            print(f"  [FAIL] P{i+1:3d}: ERROR - {e}")
            if args.verbose:
                traceback.print_exc()
            all_results.append({'page': i+1, 'status': 'EXCEPTION', 'error': str(e)})
    
    # Cross-check/backfill GRUP against the roster's theory group, if the
    # students file had one (e.g. a Moodle participants export's Grups column)
    backfill_and_validate_groups(all_results, students_df)

    # Generate outputs
    excel_path = os.path.join(args.output_dir, 'results.xlsx')
    print(f"\nWriting Excel to {excel_path}...")
    write_excel(all_results, students_df, correct_answers_by_perm, excel_path,
                args.questions, args.num_options)
    
    pdf_path = os.path.join(args.output_dir, 'annotated_review.pdf')
    print(f"Writing annotated PDF to {pdf_path}...")
    write_annotated_pdf(all_results, pdf_path, students_df=students_df)
    
    # Summary
    matched = sum(1 for r in all_results if r.get('u_number'))
    n_processed = sum(1 for r in all_results if r.get('answers'))
    print(f"\n{'='*60}")
    print(f"DONE")
    print(f"  Pages processed: {n_processed}/{len(pages)}")
    print(f"  U-numbers detected: {matched}")
    print(f"  Outputs:")
    print(f"    - {excel_path}")
    print(f"    - {pdf_path}")
    print(f"{'='*60}")


if __name__ == '__main__':
    main()
