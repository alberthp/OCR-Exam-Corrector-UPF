"""score_question() partial-credit formula, and write_excel()/
_write_group_sheet() with the input combinations most likely to appear in
real use: empty data, missing roster, missing answer key, and permutation/
option-count boundaries.
"""

import openpyxl
import pytest

import omr_correct as omr


# ----- score_question -----

def test_score_all_correct_marked_is_full_credit():
    assert omr.score_question({"A", "B"}, {"A", "B"}, num_options=4) == 1.0


def test_score_nothing_marked_is_zero():
    assert omr.score_question(set(), {"A"}, num_options=4) == 0.0


def test_score_partial_credit():
    # 1 correct out of 2 correct options, no wrong marks: +1/2
    assert omr.score_question({"A"}, {"A", "B"}, num_options=4) == pytest.approx(0.5)


def test_score_wrong_mark_is_penalized():
    # 1 correct (weight 1/1) minus 1 wrong (weight 1/(4-1)) = 1 - 1/3
    score = omr.score_question({"A", "C"}, {"A"}, num_options=4)
    assert score == pytest.approx(1 - 1 / 3)


def test_score_never_goes_negative():
    score = omr.score_question({"B", "C", "D"}, {"A"}, num_options=4)
    assert score == 0.0


def test_score_marking_every_option_cancels_to_zero():
    """By construction: marking all N options gets every correct-option
    credit and every wrong-option penalty, which cancel exactly -- this is
    what stops a student from gaming an unknown question."""
    score = omr.score_question({"A", "B", "C", "D"}, {"A", "B"}, num_options=4)
    assert score == 0.0


def test_score_good_equals_zero_returns_zero():
    # An answer-key row with no correct option at all is not a valid question.
    assert omr.score_question({"A"}, set(), num_options=4) == 0.0


def test_score_good_equals_num_options_returns_zero():
    """Documents a real (Medium-severity, not fixed) finding: an answer
    key row marking every option correct always scores 0, since there's no
    wrong option left to penalize -- yet callers still count this question
    toward max_score, silently costing every student a point. See
    RELEASE_REVIEW_v1.4.md item #11."""
    assert omr.score_question({"A"}, {"A", "B", "C", "D"}, num_options=4) == 0.0


def test_score_mark_outside_valid_option_range_is_harmless():
    # A garbage mark (never a real option letter) simply can't match
    # correct_answers and is never penalized against, since penalty
    # weighting is keyed off num_options, not the mark itself.
    assert omr.score_question({"Z"}, {"A"}, num_options=4) == 0.0


# ----- write_excel: empty / missing-input robustness -----

def test_write_excel_empty_results(tmp_path):
    out = tmp_path / "out.xlsx"
    omr.write_excel([], None, {"1": {1: {"A"}}}, str(out), num_questions=1, num_options=4)
    wb = openpyxl.load_workbook(out)
    assert "Summary" in wb.sheetnames


def test_write_excel_no_students_df(tmp_path, sample_result, sample_correct_answers):
    out = tmp_path / "out.xlsx"
    del sample_result["_corrected"]  # not needed for write_excel
    omr.write_excel([sample_result], None, sample_correct_answers, str(out),
                     num_questions=2, num_options=4)
    wb = openpyxl.load_workbook(out)
    ws = wb["Perm 1"]
    assert ws.cell(row=4, column=3).value == "000001"  # U_Number column


def test_write_excel_no_answer_key(tmp_path, sample_result):
    out = tmp_path / "out.xlsx"
    del sample_result["_corrected"]
    omr.write_excel([sample_result], None, {}, str(out), num_questions=2, num_options=4)
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["No_Perm_Detected", "T1", "T2", "Summary"]


def test_write_excel_num_questions_zero(tmp_path, sample_result):
    out = tmp_path / "out.xlsx"
    del sample_result["_corrected"]
    omr.write_excel([sample_result], None, {"1": {}}, str(out), num_questions=0, num_options=4)
    wb = openpyxl.load_workbook(out)
    assert "Perm 1" in wb.sheetnames  # doesn't crash; just no question columns


def test_write_excel_orphan_permutation_gets_empty_sheet(tmp_path, sample_result):
    """An answer key can legitimately list a permutation no scanned page
    used (e.g. a make-up exam version); it should get its own (empty)
    sheet, not be dropped or cause an error."""
    out = tmp_path / "out.xlsx"
    del sample_result["_corrected"]
    correct = {"1": {1: {"A"}}, "99": {1: {"B"}}}
    omr.write_excel([sample_result], None, correct, str(out), num_questions=1, num_options=4)
    wb = openpyxl.load_workbook(out)
    assert "Perm 99" in wb.sheetnames
    ws = wb["Perm 99"]
    assert ws.cell(row=4, column=1).value is None  # no data rows


# ----- num_options boundaries -----

@pytest.mark.parametrize("num_options", [2, 10])
def test_write_excel_num_options_boundaries(tmp_path, sample_result, num_options):
    out = tmp_path / "out.xlsx"
    del sample_result["_corrected"]
    opt_letters = omr.OPTION_LABELS[:num_options]
    correct = {"1": {1: {opt_letters[0]}}}
    sample_result["answers"] = {1: {"marks": {opt_letters[0]}}}
    omr.write_excel([sample_result], None, correct, str(out), num_questions=1, num_options=num_options)
    wb = openpyxl.load_workbook(out)
    ws = wb["Perm 1"]
    header_row = [c.value for c in ws[2]]
    assert header_row[-1] == "Score"
    assert header_row.count("Score") == 1
